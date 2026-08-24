#!/usr/bin/env python3
"""
Mailinglisten-Abgleich Core-Logik
==================================
Reine Datenlogik ohne UI-Abhängigkeiten — wird von der GUI und vom
kopflosen `probelauf.py` gleichermaßen benutzt.

Aufgabe: den Access-Adressstamm des Verlags gegen die Lexware-Exporte
abgleichen. Wer im Berichtszeitraum gekauft hat, bekommt in Access ein
aktuelles `Bestelldatum`; echte Neukunden werden angelegt.

Die beiden Systeme haben **keinen gemeinsamen Schlüssel** — in Access steht
nirgends eine Lexware-Kd.-Nr. Deshalb muss über Name und Adresse gerechnet
werden, und deshalb schreibt das Tool die Kd.-Nr. beim ersten Lauf in das
Access-Feld `Lexware-Kd-Nr` zurück: ab dem Folgejahr ist der Abgleich exakt
statt geraten.
"""

from __future__ import annotations

import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------
# Daten-Ordner (Config, Ausgabe)
# ---------------------------------------------------------------------

def _base_dir() -> Path:
    """Ordner neben der .exe (PyInstaller-Build) bzw. neben dem Tool-Code."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


# Laufzeitdateien liegen direkt neben der .exe — bewusst KEIN data-Unterordner,
# damit der Anwender sie sofort findet.
APP_DIR = _base_dir()
CONFIG_PFAD = APP_DIR / "config.json"


# ---------------------------------------------------------------------
# Schwellenwerte der Einstufung
# ---------------------------------------------------------------------
# Diese drei Zahlen entscheiden, wie viel Handarbeit übrig bleibt. Sie sind
# ein begründeter Startwert, kein Messergebnis — nach dem ersten echten
# Durchgang anhand von protokoll.xlsx nachziehen.

SCHWELLE_SICHER = 92.0    # ab hier gilt ein Treffer als eindeutig
SCHWELLE_UNKLAR = 75.0    # so gut darf der Zweitbeste höchstens sein
SCHWELLE_MINDEST = 60.0   # darunter wird gar nicht erst vorgelegt

# ... es sei denn, der Beste liegt so weit vorn UND der Zweitplatzierte ist
# erkennbar jemand anderes (abweichender Vorname unter derselben Anschrift —
# ein Haushaltsmitglied). Dann ist der exakte Treffer der richtige.
#
# Trägt der Zweite dagegen denselben Namen, hilft auch ein großer Abstand
# nicht: dann steht dieselbe Person zweimal in Access, meist mit privater und
# dienstlicher Anschrift. Welche gemeint ist — oder ob beide Post bekommen
# sollen — kann nur jemand entscheiden, der die Leute kennt.
SCHWELLE_ABSTAND = 15.0

# Sollen abweichende Adressfelder aus Lexware standardmäßig übernommen werden?
# Ja — Lexware ist das System, in dem täglich gearbeitet wird, und eine
# Adresse, unter der 2025/2026 tatsächlich geliefert wurde, ist aktueller als
# der Access-Stand. Der Bediener kann jede Übernahme einzeln abwählen.
UEBERNAHME_VORGABE = True

# Fälle, wie sie die GUI in Reitern zeigt
FALL_NEU = "neu"
FALL_AKTUALISIEREN = "aktualisieren"
FALL_UNKLAR = "unklar"
FALL_OHNE_AUFTRAG = "ohne_auftrag"


# ---------------------------------------------------------------------
# Einlesen
# ---------------------------------------------------------------------
# Die Exporte sind unsauber: der Lexware-Kundenexport trägt in Zeile 1–2 einen
# fremden Access-Header samt Beispielzeile (jemand hat das Zielformat
# hineinkopiert), der echte Header steht erst in Zeile 5. Die Aufträge-Datei
# hat eine Titelzeile über dem Header. Auf feste Zeilennummern ist also kein
# Verlass — die Kopfzeile wird gesucht.

ACCESS_PFLICHT = {"ID", "Name", "PLZ", "Bestelldatum"}
LEXWARE_PFLICHT = {"Kd.-Nr", "Name", "Plz"}
AUFTRAEGE_PFLICHT = {"Datum", "Kd.-Nr.", "Belegnr.", "Art", "Gesamt"}

# Belegarten, die einen Kauf belegen. RG ist die Rechnung, SR die
# Sammelrechnung (Barsortimente bekommen eine Monatsrechnung statt vieler
# Einzelrechnungen). Nicht dabei: LS (Lieferschein — Rezensions- und
# Autorenexemplare gehen so raus), GS (Gutschrift), ST (Storno),
# PR (Proforma), AB (Auftragsbestätigung), AG (Angebot).
KAUF_ARTEN = {"RG", "SR"}


def finde_kopfzeile(zeilen: list[tuple], pflichtspalten: set[str]) -> int:
    """Index der ersten Zeile, die alle Pflichtspalten enthält (0-basiert)."""
    for i, zeile in enumerate(zeilen):
        vorhanden = {str(z).strip() for z in zeile if z is not None}
        if pflichtspalten <= vorhanden:
            return i
    raise ValueError(
        "Kopfzeile nicht gefunden — erwartet wurden die Spalten "
        + ", ".join(sorted(pflichtspalten))
    )


def lade_tabelle(pfad, pflichtspalten: set[str]) -> tuple[list[str], list[dict]]:
    """Liest ein Blatt als Liste von dicts. Gibt (Spaltennamen, Zeilen) zurück.

    Die Spaltennamen werden mitgegeben, weil die Ausgabe später exakt die
    Reihenfolge des Access-Exports treffen muss — ein verschobener Anfüge-
    Import scheitert in Access wortlos.
    """
    wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        zeilen = [tuple(z) for z in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    kopf_i = finde_kopfzeile(zeilen, pflichtspalten)
    spalten = [
        (str(z).strip() if z is not None else f"_leer{i}")
        for i, z in enumerate(zeilen[kopf_i])
    ]

    daten = []
    for zeile in zeilen[kopf_i + 1:]:
        if not any(z not in (None, "") for z in zeile):
            continue
        satz = {}
        for i, name in enumerate(spalten):
            satz[name] = zeile[i] if i < len(zeile) else None
        daten.append(satz)
    return spalten, daten


def lade_access(pfad) -> tuple[list[str], list[dict]]:
    return lade_tabelle(pfad, ACCESS_PFLICHT)


def lade_lexware_kunden(pfad) -> tuple[list[str], list[dict]]:
    spalten, daten = lade_tabelle(pfad, LEXWARE_PFLICHT)
    # Der eingeschleppte Access-Beispielsatz oberhalb des echten Headers wird
    # von finde_kopfzeile bereits übersprungen; hier bleibt nur noch, Sätze
    # ohne Kd.-Nr. auszusortieren.
    daten = [s for s in daten if str(s.get("Kd.-Nr") or "").strip()]
    return spalten, daten


def lade_auftraege(pfade) -> list[dict]:
    """Liest beliebig viele Aufträge-Dateien in eine gemeinsame Liste.

    Mehrere Dateien sind der Normalfall, kein Sonderfall: der Kundenexport
    reicht bis zum letzten Mailing zurück, ein Aufträge-Export umfasst aber
    immer nur ein Kalenderjahr. Wer nur das laufende Jahr lädt, verliert für
    alle im Vorjahr angelegten Kunden das Bestelljahr — siehe
    pruefe_zeitraeume().
    """
    alle = []
    for pfad in pfade:
        _, daten = lade_tabelle(pfad, AUFTRAEGE_PFLICHT)
        for satz in daten:
            satz["_quelle"] = Path(pfad).name
        alle.extend(daten)
    return alle


def auftrags_spalten(auftraege: list[dict]) -> set:
    """Welche Spalten in den geladenen Aufträgen überhaupt vorkommen.

    Über ALLE Zeilen, nicht über eine Stichprobe: die Dateien haben oft
    unterschiedliche Spaltensätze, und die reichhaltigere kann hinten liegen.
    """
    spalten = set()
    for satz in auftraege:
        spalten |= set(satz)
    return spalten


# Adressfelder, wie die Auftragsliste sie nennt, auf die Namen des
# Kundenexports abgebildet. Beides sind Lexware-Ansichten derselben Daten,
# aber die Spalten heißen nicht gleich — die Hausnummer etwa heißt in der
# Kundenliste "Haus Nr." und in der Auftragsliste "Hausnummer". Je Zielfeld
# stehen deshalb mehrere mögliche Quellnamen; der erste vorhandene gewinnt.
_AUFTRAG_ZU_KUNDE = {
    "Kd.-Nr": ["Kd.-Nr.", "Kd.-Nr"],
    "Name": ["Name"],
    "Vorname": ["Vorname"],
    "Firma": ["Firma"],
    "Zusatz": ["Zusatz"],
    "Anrede": ["Anrede"],
    "Straße": ["Straße"],
    "Haus Nr.": ["Haus Nr.", "Hausnummer", "Haus-Nr."],
    "Land": ["Land"],
    "Plz": ["Plz", "PLZ"],
    "Ort": ["Ort"],
    "E-Mail": ["E-Mail", "eMail"],
    "Tel1": ["Tel1", "Telefon"],
    "Mobil": ["Mobil"],
    "Fax": ["Fax", "Telefax"],
    "Matchcode": ["Matchcode"],
    "Kundengruppe": ["Kundengruppe", "Kd.Gr."],
    "Branche": ["Branche"],
}


def ergaenze_kunden_aus_auftraegen(kunden: list[dict],
                                   auftraege: list[dict]) -> list[dict]:
    """Kunden, die nur in den Aufträgen vorkommen, aus deren Adressspalten
    nachbilden.

    Der Kundenexport reicht nur bis zum letzten Mailing zurück; wer davor
    angelegt wurde und jetzt wieder bestellt, steht ausschließlich in den
    Aufträgen. Trägt die Auftragsliste Adressspalten (das ist eine Frage der
    Listeneinstellungen in Lexware), lassen sich diese Bestandskunden daraus
    rekonstruieren und ganz normal abgleichen.

    Fehlen die Adressspalten, passiert hier schlicht nichts — dann bleibt der
    Kundenvollexport der einzige Weg.
    """
    bekannt = {str(k.get("Kd.-Nr") or "").strip() for k in kunden}
    ergaenzt: dict[str, dict] = {}

    for satz in auftraege:
        kdnr = str(satz.get("Kd.-Nr.") or "").strip()
        if not kdnr or kdnr in bekannt:
            continue

        neu = {}
        for kundenfeld, quellnamen in _AUFTRAG_ZU_KUNDE.items():
            for quelle in quellnamen:
                wert = satz.get(quelle)
                if wert not in (None, ""):
                    neu[kundenfeld] = wert
                    break

        # Die Auftragsliste führt PLZ und Ort oft in einem Feld.
        if not neu.get("Plz") and satz.get("Plz, Ort"):
            plz, ort = teile_plz_ort(satz.get("Plz, Ort"))
            if plz:
                neu["Plz"] = plz
            if ort:
                neu["Ort"] = ort

        # Ohne Ortsangabe ist ein Satz für den Abgleich wertlos — dann lieber
        # gar keinen Kandidaten erzeugen als einen unbrauchbaren.
        if not neu.get("Plz") or not (neu.get("Name") or neu.get("Firma")):
            continue

        neu["Kd.-Nr"] = kdnr
        neu["_aus_auftrag"] = True
        # Der jüngste Beleg gewinnt: spätere Zeilen überschreiben frühere.
        ergaenzt[kdnr] = neu

    return kunden + list(ergaenzt.values())


# ---------------------------------------------------------------------
# Normalisieren
# ---------------------------------------------------------------------

def norm_text(wert) -> str:
    """Kleinschreibung, Umlaute ausgeschrieben, Satzzeichen und Mehrfach-
    leerzeichen weg. Basis für jeden Vergleich."""
    if wert is None:
        return ""
    s = str(wert).strip().lower()
    # Vor der Unicode-Zerlegung ersetzen, sonst wird aus "ü" ein "u" statt "ue"
    # und "Müller" träfe nicht mehr auf "Mueller".
    for alt, neu in (("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss")):
        s = s.replace(alt, neu)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_strasse(wert) -> str:
    """Wie norm_text, zusätzlich alle Leerzeichen entfernt und die Endung
    Straße/Str. vereinheitlicht.

    Erst zusammenziehen, dann die Endung ersetzen — deutsche Straßennamen
    werden meist zusammengeschrieben ("Hasenpfuhlstraße"), und eine Ersetzung
    auf Wortgrenzen fände dort gar nichts. Genau daran scheiterte der
    Vergleich "Hasenpfuhlstr." gegen "Hasenpfuhlstraße".

    Nur `strasse` und `str` am Ende, ausdrücklich nicht `st`: sonst würde aus
    "Am Forst" ein "Am Forstr".
    """
    s = norm_text(wert).replace(" ", "")
    return re.sub(r"(strasse|str)$", "str", s)


def norm_plz(wert) -> str:
    if wert is None:
        return ""
    return re.sub(r"\D", "", str(wert)).strip()


def norm_mail(wert) -> str:
    if wert is None:
        return ""
    return str(wert).strip().lower()


def norm_hausnr(wert) -> str:
    """Hausnummer auf Ziffern und angehängten Buchstaben reduzieren
    ("36 a" und "36a" sind dieselbe Tür)."""
    if wert is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(wert).strip().lower())


# Hausnummer am Ende eines Straßenfeldes: "Untere Hauptstr. 54", "Am Berg 3a",
# "Wilhelmstr. 14-18". Der Straßenteil muss dabei übrig bleiben und lang genug
# sein — sonst zerlegte die Regel Mannheimer Quadrate wie "C 5" in die Straße
# "C" und die Hausnummer "5", und die halbe Innenstadt fände sich nicht mehr.
# Höchstens vier Ziffern je Teil: „14-18" ist eine Hausnummernspanne,
# „60191-1234" eine amerikanische ZIP+4, die zufällig am Ende einer Zeile
# steht. Ohne die Grenze wurde aus „Wood Dale IL 60191-1234" die Straße
# „Wood Dale IL" mit der Hausnummer „60191-1234".
_HAUSNR_AM_ENDE = re.compile(
    r"^(.{3,}?)[\s,]+(\d{1,4}\s*[a-zA-Z]?(?:\s*[-/]\s*\d{1,4}\s*[a-zA-Z]?)?)$")


def teile_strasse(strasse, hausnr="") -> tuple[str, str]:
    """(Straße, Hausnummer) — auch wenn beides in einem Feld steht.

    Lexware führt die Hausnummer mal getrennt, mal am Ende der Straße. Ohne
    diese Trennung gilt „Untere Hauptstr. 54" gegen „Untere Hauptstr." + „54"
    als Abweichung, obwohl es dieselbe Anschrift ist.
    """
    s = str(strasse or "").strip()
    n = str(hausnr or "").strip()
    if n or not s:
        return s, n
    treffer = _HAUSNR_AM_ENDE.match(s)
    if treffer:
        return treffer.group(1).strip(), re.sub(r"\s+", "",
                                                treffer.group(2))
    return s, n


def teile_plz_ort(wert) -> tuple[str, str]:
    """Die Aufträge-Datei führt PLZ und Ort in EINEM Feld ("36244 Bad
    Hersfeld"). Auslandsadressen haben teils Buchstabenpräfixe (CH-8000)."""
    s = str(wert or "").strip()
    # Der Bindestrich gehört bei polnischen („59-241") und amerikanischen
    # („60191-1234") Postleitzahlen zur Zahl und darf sie nicht beenden.
    # Zwei bis fünf Ziffern, weil polnische Codes „59-241" vorn nur zwei
    # haben. Das Feld enthält immer PLZ + Ort, deshalb kann die Zahl am
    # Anfang nichts anderes sein.
    treffer = re.match(r"^\s*([A-Za-z]{1,3}-)?(\d{2,5}(?:-\d{2,4})?)\s+(.*)$", s)
    if treffer:
        return treffer.group(2), treffer.group(3).strip()
    return "", s


def aehnlich(a: str, b: str) -> float:
    """Ähnlichkeit zweier bereits normalisierter Zeichenketten, 0.0–1.0."""
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    return SequenceMatcher(None, a, b).ratio()


def aehnlich_worte(a: str, b: str) -> float:
    """Wortweise Ähnlichkeit — gleichgültig gegen die Reihenfolge."""
    wa, wb = set(a.split()), set(b.split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def aehnlich_org(a: str, b: str) -> float:
    """Für Firmen- und Institutionsnamen.

    Zeichenweise UND wortweise vergleichen, der bessere Wert gilt. Firmen
    schreiben sich in beiden Systemen selten gleich, und „Blätterwald
    Buchhandlung" gegen „Buchhandlung Blätterwald" ist dieselbe Buchhandlung —
    zeichenweise aber ein schlechter Treffer, weil die Reihenfolge zählt.
    """
    return max(aehnlich(a, b), aehnlich_worte(a, b))


# ---------------------------------------------------------------------
# Bestelljahre aus den Aufträgen
# ---------------------------------------------------------------------

def _jahr(wert) -> int | None:
    if wert is None:
        return None
    if hasattr(wert, "year"):
        return int(wert.year)
    treffer = re.search(r"(19|20)\d{2}", str(wert))
    return int(treffer.group(0)) if treffer else None


def _betrag(wert) -> float:
    if isinstance(wert, (int, float)):
        return float(wert)
    try:
        return float(str(wert).replace(".", "").replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


@dataclass
class Auftragslage:
    """Was die Aufträge über einen Kunden hergeben."""
    kaufjahr: dict[str, int] = field(default_factory=dict)
    hat_beleg: set[str] = field(default_factory=set)
    fruehestes_datum: object = None
    spaetestes_datum: object = None


def bestelljahre(auftraege: list[dict]) -> Auftragslage:
    """Je Kd.-Nr. das jüngste Jahr mit einem echten Kauf.

    Als Kauf zählt nur eine Rechnung (`Art` = RG) mit Umsatz > 0. Rezensions-
    und Autorenexemplare gehen auf einem Lieferschein oder mit Betrag 0 raus —
    das ist keine Bestellung. Die Betroffenen werden trotzdem angelegt, nur
    eben ohne `Bestelldatum`; ihr Brief kommt über das Merkmal (Presse, Autor),
    nicht über den Kauf. Genau so wurde es bisher auch von Hand gemacht.

    `hat_beleg` sammelt dagegen JEDEN Beleg — das unterscheidet später
    "Freiexemplar" von "Aufträge-Datei fehlt".
    """
    lage = Auftragslage()
    for satz in auftraege:
        kdnr = str(satz.get("Kd.-Nr.") or "").strip()
        if not kdnr:
            continue
        jahr = _jahr(satz.get("Datum"))
        lage.hat_beleg.add(kdnr)

        datum = satz.get("Datum")
        if datum is not None and hasattr(datum, "year"):
            if lage.fruehestes_datum is None or datum < lage.fruehestes_datum:
                lage.fruehestes_datum = datum
            if lage.spaetestes_datum is None or datum > lage.spaetestes_datum:
                lage.spaetestes_datum = datum

        art = str(satz.get("Art") or "").strip().upper()
        if jahr and art in KAUF_ARTEN and _betrag(satz.get("Gesamt")) > 0:
            if kdnr not in lage.kaufjahr or jahr > lage.kaufjahr[kdnr]:
                lage.kaufjahr[kdnr] = jahr
    return lage


# ---------------------------------------------------------------------
# Plausibilitätsprüfung der Eingaben
# ---------------------------------------------------------------------

@dataclass
class Zeitraumbefund:
    verdaechtig: bool
    ohne_beleg: int
    gesamt: int
    bruch_kdnr: str | None
    anteil_davor: float
    anteil_danach: float
    fruehestes_datum: object
    text: str


def pruefe_zeitraeume(kunden: list[dict], lage: Auftragslage,
                      block: int = 25) -> Zeitraumbefund:
    """Erkennt, ob eine Aufträge-Datei fehlt.

    Diese Falle stellt sich jedes Jahr neu und ist von außen unsichtbar: es
    fehlt nichts erkennbar, es sieht nur nach "Kunde ohne Bestellung" aus.
    Da in Lexware ein Kundensatz erst entsteht, wenn eine Rechnung geschrieben
    wird, ist ein Kunde ganz ohne Beleg aber ein Widerspruch.

    Die Signatur ist eindeutig: über die aufsteigende Kd.-Nr. bricht der
    Anteil der Belegloser an einer Stelle von ~95 % auf ~0 % ein — dort endet
    der Zeitraum, den die geladenen Aufträge abdecken.
    """
    kdnrs = sorted(
        (str(k.get("Kd.-Nr") or "").strip() for k in kunden),
        key=lambda s: (len(s), s),
    )
    kdnrs = [k for k in kdnrs if k]
    gesamt = len(kdnrs)
    ohne = [k for k in kdnrs if k not in lage.hat_beleg]

    if not gesamt or not ohne:
        return Zeitraumbefund(False, len(ohne), gesamt, None, 0.0, 0.0,
                              lage.fruehestes_datum,
                              "Alle Lexware-Kunden haben einen Beleg.")

    # Gleitend über Blöcke von je `block` Kunden den Anteil ohne Beleg bilden
    # und die größte Stufe suchen.
    anteile = []
    for i in range(0, gesamt, block):
        teil = kdnrs[i:i + block]
        anteile.append(sum(1 for k in teil if k not in lage.hat_beleg) / len(teil))

    bester_sprung, bester_i = 0.0, None
    for i in range(1, len(anteile)):
        davor = sum(anteile[:i]) / i
        danach = sum(anteile[i:]) / (len(anteile) - i)
        if davor - danach > bester_sprung:
            bester_sprung, bester_i = davor - danach, i

    anteil_gesamt = len(ohne) / gesamt
    if bester_i is None or bester_sprung < 0.5:
        # Kein klarer Bruch — vermutlich nur verstreute Freiexemplare.
        return Zeitraumbefund(
            anteil_gesamt > 0.15, len(ohne), gesamt, None,
            0.0, 0.0, lage.fruehestes_datum,
            f"{len(ohne)} von {gesamt} Lexware-Kunden ohne jeden Beleg "
            f"({anteil_gesamt:.0%}), aber ohne klaren Bruch im Nummernkreis.",
        )

    bruch = kdnrs[min(bester_i * block, gesamt - 1)]
    davor = sum(anteile[:bester_i]) / bester_i
    danach = sum(anteile[bester_i:]) / (len(anteile) - bester_i)
    datum_text = (str(lage.fruehestes_datum)[:10]
                  if lage.fruehestes_datum is not None else "unbekannt")
    return Zeitraumbefund(
        True, len(ohne), gesamt, bruch, davor, danach, lage.fruehestes_datum,
        f"{len(ohne)} von {gesamt} Lexware-Kunden haben in den geladenen "
        f"Aufträge-Dateien KEINEN Beleg. Unterhalb Kd.-Nr. {bruch} sind es "
        f"{davor:.0%}, darüber nur {danach:.0%}. Das ist die Signatur einer "
        f"fehlenden Aufträge-Datei: der früheste geladene Auftrag stammt vom "
        f"{datum_text}, der Kundenexport reicht weiter zurück. "
        f"Bitte die Aufträge des Vorjahres nachladen.",
    )


# Felder, die einen Satz unabhängig vom Kaufdatum im Mailing halten.
MERKMALSFELDER = ("Autor", "Kommunen", "Bürgermeister", "Verein", "Schule",
                  "Medium", "Presse/ZS", "Veranstalter", "VIP/W/K/X")

# `K` ist die einzige Kategorie, die am Kauf hängt — bei ihr halten sich
# Sätze mit und ohne aktuelles Bestelldatum die Waage (5 278 zu 2 766). Alle
# anderen stehen zu 91–100 % OHNE aktuellen Kauf im Mailing-Filter, sind also
# über ihre Kategorie dauerhaft dabei: Buchhandel, Presse, Wissenschaft,
# Vereine, VIP. Für sie ändert ein neues Bestelldatum nichts daran, ob ein
# Brief kommt — nur daran, was in der Datenbank steht.
#
# Abgelesen am Export selbst, denn der IST die Mailing-Abfrage.
KATEGORIE_KAUFABHAENGIG = {"K", ""}


def dauerhaft_im_mailing(acc: dict) -> str:
    """Kategorie, wenn der Satz unabhängig vom Bestelldatum Post bekommt."""
    kat = str(acc.get("VIP/W/K/X") or "").strip()
    return "" if kat.upper() in {k.upper() for k in KATEGORIE_KAUFABHAENGIG} \
        else kat


def pruefe_access_vollstaendigkeit(access: list[dict], laufjahr: int,
                                   jahre: int = 3) -> str | None:
    """Erkennt, ob statt der Kundentabelle nur die Mailing-Abfrage exportiert
    wurde.

    Der Kronzeuge sind **Kunden der Kategorie K mit altem Bestelldatum**. Nur
    bei ihnen hängt der Brief am Kaufjahr; wer 2015 zuletzt gekauft hat, wäre
    aus einer Mailing-Abfrage „letzte drei Jahre" herausgefallen. Stehen
    trotzdem viele davon im Export, ist es die volle Tabelle.

    Ein früherer Versuch prüfte, ob Sätze ohne aktuelles Bestelldatum ein
    Merkmal tragen — und war doppelt untauglich: `VIP/W/K/X` ist bei praktisch
    jedem Satz gefüllt, und ein Bestelldatum führt der Verlag ohnehin nur bei
    K-Kunden. Die Meldung erschien deshalb immer.
    """
    if not access:
        return None
    grenze = laufjahr - jahre
    k_alt = k_gesamt = 0
    for satz in access:
        if str(satz.get("VIP/W/K/X") or "").strip().upper() != "K":
            continue
        k_gesamt += 1
        jahr = satz.get("Bestelldatum")
        if not isinstance(jahr, int) or jahr < grenze:
            k_alt += 1

    if k_gesamt < 50:
        return None                        # zu wenig Material für ein Urteil
    anteil = k_alt / k_gesamt
    if anteil >= 0.10:
        return None                        # klar die volle Tabelle

    return (
        f"Die Access-Datei könnte die Mailing-Abfrage sein statt der vollen "
        f"Kundentabelle: von {k_gesamt} Sätzen der Kategorie K haben nur "
        f"{k_alt} ein Bestelldatum, das älter als {grenze} ist ({anteil:.1%}). "
        f"Bei K hängt der Brief am Kaufjahr — eine Abfrage „letzte {jahre} "
        f"Jahre“ enthielte solche Sätze gar nicht, die volle Tabelle dagegen "
        f"reichlich. Folge: Altkunden würden nicht gefunden und ein ZWEITES "
        f"Mal angelegt. Bitte prüfen, ob wirklich die ganze Tabelle "
        f"exportiert wurde.")


def pruefe_spalten(access_spalten: list[str], kunden_spalten: list[str],
                   auftrags_spalten: set) -> list[str]:
    """Meldet Spalten, die zwar nicht Pflicht sind, aber die Trefferquote
    deutlich heben — oder deren Fehlen ganze Fallgruppen unmöglich macht.

    Der Sinn: der Bediener soll beim Laden erfahren, dass sein Export etwas
    Wichtiges nicht enthält, statt es am schlechten Ergebnis zu erraten. Das
    Auswählen der Spalten passiert in Lexware unter Ansicht ->
    Listeneinstellungen, und was dort einmal fehlte, fehlt jedes Jahr wieder.
    """
    mangel = []

    if "Lexware-Kd-Nr" not in access_spalten:
        mangel.append(
            "Access: Feld 'Lexware-Kd-Nr' fehlt. Ohne dieses Feld bleibt der "
            "Abgleich auch im Folgejahr geraten statt exakt — anlegen "
            "(Text, 20) und neu exportieren lohnt sich einmalig sehr.")

    for feld, wozu in (
            ("Haus Nr.", "Hausnummern trennen Nachbarn"),
            ("E-Mail", "die E-Mail ist das stärkste Einzelmerkmal"),
            ("Plz", "ohne PLZ ist kein Blocken möglich")):
        if feld not in kunden_spalten:
            mangel.append(f"Lexware-Kunden: Spalte '{feld}' fehlt — {wozu}.")

    # Adressspalten in den Aufträgen sind kein Muss, aber ohne sie bleiben
    # Bestandskunden außen vor, sofern der Kundenexport sie nicht enthält.
    hat_adresse = bool(auftrags_spalten & {"Straße", "Plz", "Plz, Ort"})
    if not hat_adresse:
        mangel.append(
            "Aufträge: keine Adressspalten. Bestandskunden, die nicht im "
            "Kundenexport stehen, lassen sich dann nicht zuordnen — entweder "
            "den Kundenexport ohne Einschränkung ziehen oder in den "
            "Listeneinstellungen der Auftragsliste Name, Firma, Plz, Ort, "
            "Straße und Hausnummer dazuschalten.")
    else:
        if not (auftrags_spalten & {"Haus Nr.", "Hausnummer", "Haus-Nr."}):
            mangel.append(
                "Aufträge: keine Hausnummer — Bestandskunden werden nur über "
                "die Straße erkannt, nicht über die genaue Anschrift.")
        if not (auftrags_spalten & {"E-Mail", "eMail"}):
            mangel.append(
                "Aufträge: keine E-Mail. Die Auftragsliste bietet sie meist "
                "gar nicht an (nur die eRechnungs-Adresse) — für Bestands"
                "kunden entfällt damit das stärkste Einzelmerkmal. Wer das "
                "braucht, kommt am Kundenvollexport nicht vorbei.")
    return mangel


# ---------------------------------------------------------------------
# Abgleich
# ---------------------------------------------------------------------

@dataclass
class Kandidat:
    access: dict
    punkte: float
    begruendung: str
    gesperrt: bool = False        # verstorben/verzogen — nie automatisch
    sperrgrund: str = ""
    vorname_konflikt: bool = False  # gleicher Haushalt, andere Person
    # IDs gleichlautender Access-Sätze, die zu diesem zusammengefasst wurden
    dubletten: list = field(default_factory=list)


# Was mit einem Fall geschehen soll. Die GUI setzt das um; `stufe_ein` legt
# nur den Vorschlag fest.
AKTION_NEU = "neu anlegen"
AKTION_AKTUALISIEREN = "aktualisieren"
AKTION_NICHTS = "ignorieren"

# Felder, die beim Zusammenlegen mehrere Werte vertragen, statt dass einer
# gewinnt. Zwei Adressen desselben Menschen haben oft zwei E-Mail-Adressen,
# und beide sind richtig — wegzuwerfen wäre schade.
MEHRWERTIG = ("eMail", "Telefon1", "Telefon2", "Telefon3", "Telefax")
TRENNER_MEHRWERTIG = ", "
TRENNER_BEMERKUNG = " · "


@dataclass
class Zuordnung:
    lexware: dict
    fall: str
    kandidaten: list[Kandidat] = field(default_factory=list)
    bestelljahr: int | None = None
    # Sachliche Anmerkungen zum Satz (Freiexemplar, kein Beleg …)
    hinweise: list[str] = field(default_factory=list)
    # Warum der Fall NICHT automatisch entschieden werden konnte. Getrennt
    # gehalten, weil beides in der Liste sonst durcheinandergeht: dass jemand
    # ein Rezensionsexemplar bekommen hat, ist kein Grund für Unklarheit.
    unklar_grund: list[str] = field(default_factory=list)

    # Wie der Abgleich den Fall ursprünglich eingestuft hat — damit sich jede
    # Entscheidung zurücknehmen lässt, ohne den ganzen Lauf zu wiederholen.
    fall_urspruenglich: str = ""

    # Vom Bediener entschieden (Vorbelegung aus `fall`)
    aktion: str = AKTION_NEU
    ziel: dict | None = None          # der zu aktualisierende Access-Satz
    uebernehmen: set = field(default_factory=set)  # Felder aus Lexware
    aenderungen: dict = field(default_factory=dict)  # von Hand editiert
    # Access-IDs, die in `ziel` aufgehen sollen. Sie verschwinden aus der
    # Gesamttabelle; beim Weg über die Einzeldateien müssen sie von Hand
    # gelöscht werden (siehe zusammenlegen.xlsx).
    aufgeloest_in: list = field(default_factory=list)
    # Hat der Bediener diesen Fall angefasst? Nur solche werden gesichert.
    # Andernfalls konservierte entscheidungen.json den Vorschlag eines alten
    # Programmstands und machte jede Verbesserung am Abgleich wirkungslos.
    beruehrt: bool = False

    @property
    def bester(self) -> Kandidat | None:
        return self.kandidaten[0] if self.kandidaten else None


# Felder, deren Inhalt einen Access-Satz aus dem Mailing nimmt. Wer hier etwas
# stehen hat, darf nie automatisch reaktiviert werden — einen Verstorbenen
# zurück ins Mailing zu holen ist der Fehler, den niemand sehen will.
SPERRFELDER = ("Datensatz gelöscht", "prüfen")


def dubletten_kennung(acc: dict) -> tuple:
    """Ein Access-Satz, eingedampft auf das, was ihn als Adresse ausmacht.

    Zwei Sätze mit derselben Kennung sind für dieses Werkzeug ununterscheidbar
    — dieselbe Person an derselben Anschrift, doppelt erfasst.
    """
    return (norm_text(acc.get("Name")), norm_text(acc.get("Vorname")),
            norm_text(acc.get("Institution")), norm_plz(acc.get("PLZ")),
            norm_text(acc.get("Ort")), norm_strasse(acc.get("Straße")),
            norm_hausnr(acc.get("Hausnummer")), norm_mail(acc.get("eMail")))


def _gehaltvoller(acc: dict) -> tuple:
    """Sortierschlüssel, um aus gleichlautenden Sätzen den besten zu wählen:
    der mit den meisten gefüllten Feldern, dann der zuletzt geprüfte, dann der
    mit der kleineren ID — damit die Wahl bei jedem Lauf dieselbe ist."""
    gefuellt = sum(1 for w in acc.values() if w not in (None, ""))
    geprueft = acc.get("erfaßt/geprüft am")
    stempel = geprueft.toordinal() if hasattr(geprueft, "toordinal") else 0
    ident = acc.get("ID")
    return (-gefuellt, -stempel, ident if isinstance(ident, int) else 0)


def fasse_dubletten_zusammen(kandidaten: list[Kandidat]) -> list[Kandidat]:
    """Gleichlautende Access-Sätze zu einem Kandidaten verschmelzen.

    Stehen zwei identische Sätze zur Wahl, ist das keine Entscheidung, die
    jemand treffen könnte — es ist eine Dublette in Access. Sie vorzulegen
    kostet nur Zeit. Der gehaltvollere Satz gewinnt, die anderen werden am
    Kandidaten vermerkt, damit sie im Protokoll auftauchen.
    """
    gruppen: dict[tuple, list[Kandidat]] = {}
    for k in kandidaten:
        gruppen.setdefault(dubletten_kennung(k.access), []).append(k)

    ergebnis = []
    for gruppe in gruppen.values():
        if len(gruppe) == 1:
            ergebnis.append(gruppe[0])
            continue
        gruppe.sort(key=lambda k: _gehaltvoller(k.access))
        beste = gruppe[0]
        beste.dubletten = [k.access.get("ID") for k in gruppe[1:]]
        ergebnis.append(beste)

    ergebnis.sort(key=lambda k: k.punkte, reverse=True)
    return ergebnis


def baue_indizes(access: list[dict]) -> dict:
    """Blockbildung: nur wer PLZ, E-Mail oder Straße teilt, wird verglichen.

    Ohne das wären es 18 000 × 1 000 Vergleiche; so sind es je Lexware-Kunde
    ein paar Dutzend, und difflib aus der Standardbibliothek reicht völlig.
    """
    nach_plz = defaultdict(list)
    nach_mail = defaultdict(list)
    nach_strasse = defaultdict(list)
    nach_kdnr = {}

    for satz in access:
        plz = norm_plz(satz.get("PLZ"))
        if plz:
            nach_plz[plz].append(satz)
        mail = norm_mail(satz.get("eMail"))
        if mail:
            nach_mail[mail].append(satz)
        strasse = norm_strasse(satz.get("Straße"))
        if plz and strasse:
            nach_strasse[(plz, strasse)].append(satz)
        # Mehrere Nummern je Satz sind möglich, wenn Lexware denselben
        # Menschen zweimal führt (siehe verschmolzene_aktualisierungen) —
        # dann muss jede von ihnen zum Satz führen.
        for kdnr in zerlege_kdnr(satz.get("Lexware-Kd-Nr")):
            nach_kdnr[kdnr] = satz

    return {
        "plz": nach_plz,
        "mail": nach_mail,
        "strasse": nach_strasse,
        "kdnr": nach_kdnr,
    }


def _punkte(lex: dict, acc: dict,
            viele_personen: bool = False) -> tuple[float, str, bool]:
    """Gewichtete Ähnlichkeit 0–100, Begründung, Vornamenskonflikt.

    Gewichtet wird nur, was auf beiden Seiten überhaupt vorhanden ist — sonst
    würde ein fehlendes Telefon oder eine fehlende E-Mail als Unterschied
    zählen statt als Nichtwissen.
    """
    lex_name = norm_text(lex.get("Name"))
    lex_vorname = norm_text(lex.get("Vorname"))
    lex_firma = norm_text(lex.get("Firma"))
    lex_plz = norm_plz(lex.get("Plz"))
    _lex_str, _lex_hnr = teile_strasse(lex.get("Straße"), lex.get("Haus Nr."))
    lex_str = norm_strasse(_lex_str)
    lex_hnr = norm_hausnr(_lex_hnr)
    lex_mail = norm_mail(lex.get("E-Mail"))

    acc_name = norm_text(acc.get("Name"))
    acc_vorname = norm_text(acc.get("Vorname"))
    acc_inst = norm_text(acc.get("Institution"))
    acc_plz = norm_plz(acc.get("PLZ"))
    acc_str = norm_strasse(acc.get("Straße"))
    acc_hnr = norm_hausnr(acc.get("Hausnummer"))
    acc_mail = norm_mail(acc.get("eMail"))

    teile: list[tuple[float, float]] = []
    gruende: list[str] = []

    # Name bzw. Organisation. Über Kreuz vergleichen, weil in Access mal die
    # Firma im Namensfeld steht und mal in `Institution`.
    # Welche Paarung den Treffer trägt, wird mitgeführt: sonst steht bei
    # zwei verschiedenen Menschen desselben Vereins nur „Treffer zu schwach",
    # und niemand versteht, warum sie überhaupt nebeneinander stehen.
    paarungen = [
        ("Name", aehnlich(lex_name, acc_name)),            # Person / Person
        ("Einrichtung", aehnlich_org(lex_firma, acc_inst)),  # Haus / Haus
        ("Firma im Namensfeld", aehnlich_org(lex_firma, acc_name)),
        ("Name im Firmenfeld", aehnlich_org(lex_name, acc_inst)),
    ]
    woher, s_name = max(paarungen, key=lambda p: p[1])
    teile.append((40.0, s_name))
    if s_name >= 0.99:
        gruende.append(f"{woher} identisch")
    elif s_name >= 0.8:
        gruende.append(f"{woher} ähnlich")

    # Die Einrichtung EIGENS bewerten, wenn beide Seiten eine nennen. Über das
    # Maximum oben fällt sie sonst unter den Tisch: passt der Personenname
    # exakt, ist die Punktzahl schon voll, und ob dieselbe Person beim
    # Stadtarchiv oder beim Geschichtsverein sitzt, bliebe folgenlos. Genau
    # daran unterschieden sich zwei Kandidaten nicht, die sich unterscheiden.
    if lex_firma and acc_inst:
        s_inst = aehnlich_org(lex_firma, acc_inst)
        teile.append((20.0, s_inst))
        if s_inst < 0.6:
            gruende.append("andere Einrichtung")

    # Ein abweichender Vorname bei gleichem Nachnamen und gleicher Adresse ist
    # der häufigste Beinahe-Treffer überhaupt: derselbe Haushalt, eine andere
    # Person. Das darf nie automatisch zusammengeführt werden, egal wie gut
    # der Rest passt — deshalb wird es eigens vermerkt und nicht bloß
    # weggewichtet.
    vorname_konflikt = False
    if lex_vorname and acc_vorname:
        s_vor = aehnlich(lex_vorname, acc_vorname)
        teile.append((15.0, s_vor))
        if s_vor < 0.6:
            vorname_konflikt = True
            gruende.append("Vorname weicht ab")

    if lex_plz and acc_plz:
        s_plz = 1.0 if lex_plz == acc_plz else 0.0
        teile.append((10.0, s_plz))
        if not s_plz:
            gruende.append("andere PLZ")

    if lex_str and acc_str:
        s_str = aehnlich(lex_str, acc_str)
        teile.append((25.0, s_str))
        if s_str < 0.6:
            gruende.append("andere Straße")

    if lex_hnr and acc_hnr:
        teile.append((10.0, 1.0 if lex_hnr == acc_hnr else 0.0))
        if lex_hnr != acc_hnr:
            gruende.append("andere Hausnummer")

    if lex_mail and acc_mail:
        gleich = lex_mail == acc_mail
        teile.append((25.0, 1.0 if gleich else 0.0))
        gruende.append("E-Mail gleich" if gleich else "andere E-Mail")

    # Bestellt eine Einrichtung ohne Ansprechperson und stehen in Access
    # MEHRERE Menschen an dieser Anschrift, ist keiner von ihnen gemeint —
    # gemeint ist die Einrichtung. Ohne diesen Punkt bekämen alle 39
    # Beschäftigten eines Museums dieselben 100 Punkte wie das Museum selbst.
    #
    # Steht dort nur EINE Person, ist sie die Einrichtung: bei einer
    # Buchhandlung führt Access die Inhaberin, Lexware den Laden. Deshalb
    # entscheidet der Aufrufer über `viele_personen`, nicht diese Funktion —
    # sie sieht immer nur einen Kandidaten und könnte das nicht wissen.
    if viele_personen and (acc_name or acc_vorname):
        teile.append((20.0, 0.0))
        gruende.append("Access-Satz lautet auf eine von mehreren Personen")

    gewicht = sum(g for g, _ in teile)
    punkte = 100.0 * sum(g * w for g, w in teile) / gewicht if gewicht else 0.0

    # Eine gleiche E-Mail ist das stärkste Einzelmerkmal, das es hier gibt —
    # sie überlebt Umzug und Heirat. Sie darf aber keinen Vornamenskonflikt
    # überstimmen: Eheleute und Kinder teilen sich oft eine Adresse UND eine
    # Adresse, und der Nachname stimmt dann natürlich auch.
    if (lex_mail and lex_mail == acc_mail
            and s_name >= 0.7 and not vorname_konflikt):
        punkte = max(punkte, SCHWELLE_SICHER)

    return (round(punkte, 1), ", ".join(gruende) or "übereinstimmend",
            vorname_konflikt)


def finde_zuordnung(lex: dict, indizes: dict) -> list[Kandidat]:
    """Kandidaten zu einem Lexware-Kunden, absteigend nach Punkten."""
    kdnr = str(lex.get("Kd.-Nr") or "").strip()

    # Stufe 0: Access kennt die Kd.-Nr. bereits — exakt, fertig.
    treffer = indizes["kdnr"].get(kdnr)
    if treffer is not None:
        return [Kandidat(treffer, 100.0, "Lexware-Kd-Nr in Access hinterlegt")]

    # Stufe 1: Blöcke einsammeln (über id(), weil dicts nicht hashbar sind).
    roh: dict[int, dict] = {}
    plz = norm_plz(lex.get("Plz"))
    for satz in indizes["plz"].get(plz, ()):
        roh[id(satz)] = satz
    mail = norm_mail(lex.get("E-Mail"))
    if mail:
        for satz in indizes["mail"].get(mail, ()):
            roh[id(satz)] = satz
    strasse = norm_strasse(teile_strasse(lex.get("Straße"),
                                         lex.get("Haus Nr."))[0])
    if plz and strasse:
        for satz in indizes["strasse"].get((plz, strasse), ()):
            roh[id(satz)] = satz

    # Bestellt eine Einrichtung ohne Ansprechperson? Dann zählt erst einmal,
    # wie viele namentliche Personen Access an dieser Anschrift führt — davon
    # hängt ab, ob eine von ihnen gemeint sein kann.
    werte = lexware_werte(lex)
    lex_ohne_person = (bool(werte.get("Institution"))
                       and not werte.get("Name") and not werte.get("Vorname"))
    viele_personen = False
    if lex_ohne_person:
        mit_person = sum(
            1 for satz in roh.values()
            if (str(satz.get("Name") or "").strip()
                or str(satz.get("Vorname") or "").strip())
            and _punkte(lex, satz)[0] >= SCHWELLE_MINDEST)
        viele_personen = mit_person >= 2

    # Stufe 2: bewerten.
    kandidaten = []
    for satz in roh.values():
        punkte, begruendung, vorname_konflikt = _punkte(
            lex, satz, viele_personen)
        if punkte < SCHWELLE_MINDEST:
            continue
        sperrgrund = ""
        for feld in SPERRFELDER:
            wert = str(satz.get(feld) or "").strip()
            if wert:
                sperrgrund = f"{feld}: {wert}"
                break
        kandidaten.append(Kandidat(satz, punkte, begruendung,
                                   bool(sperrgrund), sperrgrund,
                                   vorname_konflikt))

    # Bei Punktgleichstand entscheidet, wie viele Felder tatsächlich
    # übereinstimmen, dann die kleinere ID. Ohne diese beiden Stufen bestimmte
    # die Einfügereihenfolge der Blöcke die Rangfolge — also der Zufall, und
    # der Kandidat mit einem widersprechenden Feld stünde mal oben, mal unten.
    def rang(k: Kandidat):
        gleiche = sum(1 for _, _, _, b in vergleiche(lex, k.access)
                      if b == GLEICH)
        ident = k.access.get("ID")
        return (-k.punkte, -gleiche, ident if isinstance(ident, int) else 0)

    kandidaten.sort(key=rang)
    return fasse_dubletten_zusammen(kandidaten)


def stufe_ein(lex: dict, kandidaten: list[Kandidat],
              bestelljahr: int | None, hat_beleg: bool) -> Zuordnung:
    """Ordnet einen Lexware-Kunden einem der vier Töpfe zu."""
    z = Zuordnung(lex, FALL_NEU, kandidaten, bestelljahr)

    if not hat_beleg:
        # Kein einziger Beleg — in Lexware ein Widerspruch, siehe
        # pruefe_zeitraeume(). Zur Sichtung, nicht zur stillen Verarbeitung:
        # `_vorbelegen` setzt hier "nichts tun". Ohne diesen Aufruf bliebe die
        # Vorbelegung des Dataclass ("neu anlegen") stehen, und bei fehlender
        # Aufträge-Datei würden Hunderte Karteileichen angelegt.
        z.fall = FALL_OHNE_AUFTRAG
        z.fall_urspruenglich = z.fall
        z.hinweise.append("kein Beleg in den geladenen Aufträge-Dateien")
        _vorbelegen(z)
        return z

    if bestelljahr is None:
        z.hinweise.append("nur Lieferschein/Nullbeleg — Freiexemplar, "
                          "kein Bestelldatum")

    if not kandidaten:
        z.fall = FALL_NEU
        return z

    bester = kandidaten[0]
    zweiter = kandidaten[1].punkte if len(kandidaten) > 1 else 0.0

    abstand = bester.punkte - zweiter
    # Ist der Zweitplatzierte erkennbar eine andere Person? Der Vornamens-
    # konflikt ist dafür das verlässlichste Zeichen: gleiche Anschrift,
    # gleicher Nachname, anderer Vorname = Haushaltsmitglied. Fehlt er, trägt
    # der zweite Satz denselben Namen — dann ist es dieselbe Person unter
    # einer weiteren Anschrift, und das gehört vorgelegt.
    zweiter_andere_person = (kandidaten[1].vorname_konflikt
                             if len(kandidaten) > 1 else True)

    # Bestellt eine Einrichtung, ist jeder Access-Satz, der auf eine Person
    # lautet, ebenfalls „jemand anderes" — auch ohne Vornamenskonflikt, den
    # es hier gar nicht geben kann, weil Lexware keinen Vornamen führt.
    werte = lexware_werte(lex)
    lex_ohne_person = (bool(werte.get("Institution"))
                       and not werte.get("Name") and not werte.get("Vorname"))
    personen_dort = sum(
        1 for k in kandidaten
        if str(k.access.get("Name") or "").strip()
        or str(k.access.get("Vorname") or "").strip())
    if lex_ohne_person and len(kandidaten) > 1:
        zweite = kandidaten[1].access
        if str(zweite.get("Name") or "").strip() or str(
                zweite.get("Vorname") or "").strip():
            zweiter_andere_person = True

    if bester.gesperrt:
        z.fall = FALL_UNKLAR
        z.unklar_grund.append(
            f"Access-Satz ist gesperrt ({bester.sperrgrund})")
    elif bester.punkte >= SCHWELLE_SICHER and bester.vorname_konflikt:
        # Ein abweichender Vorname wird nie automatisch zusammengeführt — das
        # hieße, zwei Menschen zu einem zu machen. Welcher Sachverhalt genau
        # vorliegt, hängt aber an der Anschrift, und die muss geprüft und
        # nicht behauptet werden: sonst steht "gleiche Adresse" über einem
        # Fall, bei dem Straße und Hausnummer verschieden sind.
        #
        # Unterhalb von SCHWELLE_SICHER greift dieser Zweig gar nicht mehr:
        # bei 62 Punkten ist nicht der Haushalt das Thema, sondern dass der
        # Treffer schlicht zu schwach ist.
        z.fall = FALL_UNKLAR
        gleiche_anschrift = (
            norm_plz(lex.get("Plz")) == norm_plz(bester.access.get("PLZ"))
            and norm_strasse(lex.get("Straße"))
            == norm_strasse(bester.access.get("Straße"))
            and norm_hausnr(lex.get("Haus Nr."))
            == norm_hausnr(bester.access.get("Hausnummer")))
        einrichtung = (str(lex.get("Firma") or "").strip()
                       or str(bester.access.get("Institution") or "").strip())
        if einrichtung and gleiche_anschrift:
            z.unklar_grund.append(
                "gleiche Einrichtung, andere Ansprechperson — "
                "Namen aktualisieren oder eigenen Satz anlegen?")
        elif gleiche_anschrift:
            z.unklar_grund.append(
                "gleiche Adresse, anderer Vorname — vermutlich derselbe "
                "Haushalt, also eigener Datensatz?")
        else:
            z.unklar_grund.append(
                "gleicher Nachname, anderer Vorname, andere Anschrift — "
                "vermutlich gar nicht dieselbe Person")
    elif bester.punkte >= SCHWELLE_SICHER and (
            zweiter < SCHWELLE_UNKLAR
            or (abstand >= SCHWELLE_ABSTAND and zweiter_andere_person)):
        z.fall = FALL_AKTUALISIEREN
    elif lex_ohne_person and personen_dort >= 2 and (
            str(bester.access.get("Name") or "").strip()
            or str(bester.access.get("Vorname") or "").strip()):
        # Die Einrichtung hat bestellt, in Access stehen dort aber nur
        # namentliche Mitarbeiter. Einen davon zu aktualisieren hieße, den
        # Brief an eine zufällige Person zu adressieren.
        z.fall = FALL_UNKLAR
        z.unklar_grund.append(
            f"Die Einrichtung hat bestellt, Access führt dort aber nur "
            f"{personen_dort} einzelne Personen und keinen Satz für die "
            f"Einrichtung selbst — eigenen anlegen?")
    elif bester.punkte >= SCHWELLE_MINDEST:
        z.fall = FALL_UNKLAR
        if zweiter >= SCHWELLE_UNKLAR and not zweiter_andere_person:
            z.unklar_grund.append(
                f"derselbe Name steht mehrfach in Access "
                f"({bester.punkte:.0f} gegen {zweiter:.0f}) — zweite "
                f"Anschrift derselben Person?")
        elif zweiter >= SCHWELLE_UNKLAR:
            z.unklar_grund.append(
                f"zwei fast gleich gute Kandidaten "
                f"({bester.punkte:.0f} gegen {zweiter:.0f})")
        else:
            z.unklar_grund.append(
                f"bester Treffer nur {bester.punkte:.0f} Punkte — "
                f"zu wenig für eine automatische Zuordnung")
    else:
        z.fall = FALL_NEU

    z.fall_urspruenglich = z.fall
    _vorbelegen(z)
    return z


def zuruecksetzen(z: Zuordnung) -> None:
    """Alle Entscheidungen zu einem Fall verwerfen und neu vorbelegen."""
    z.fall = z.fall_urspruenglich or z.fall
    z.ziel = None
    z.uebernehmen = set()
    z.aenderungen = {}
    _vorbelegen(z)


def aktiviere(z: Zuordnung) -> None:
    """Einen beiseitegelegten Fall doch behandeln.

    Gebraucht, wenn im Reiter „Ignoriert" oder „Ohne Auftrag" ein Haken
    gesetzt wird: der Satz soll dann so eingeordnet werden, als läge ein
    Auftrag vor — bei gutem Access-Treffer also nach „Aktualisieren", sonst
    nach „Neu anlegen". Vorher machte das Häkchen dort etwas Willkürliches,
    weil es die Aktion am angezeigten Reiter festmachte.
    """
    bester = z.bester
    if bester and not bester.gesperrt and not bester.vorname_konflikt \
            and bester.punkte >= SCHWELLE_SICHER:
        z.fall = FALL_AKTUALISIEREN
        z.aktion = AKTION_AKTUALISIEREN
        z.ziel = bester.access
        if UEBERNAHME_VORGABE:
            z.uebernehmen = {
                f for f, (alt, neu) in abweichungen(z).items()
                if uebernahme_sinnvoll(f, alt, neu, z.ziel)}
    elif bester:
        # Kandidaten da, aber keiner sicher — das ist genau der unklare Fall.
        z.fall = FALL_UNKLAR
        z.aktion = AKTION_AKTUALISIEREN if z.ziel else AKTION_NEU
    else:
        z.fall = FALL_NEU
        z.aktion = AKTION_NEU
        z.ziel = None


def _vorbelegen(z: Zuordnung) -> None:
    """Vorschlag für die Aktion. Die GUI darf jederzeit anders entscheiden.

    Unklare Fälle bekommen bewusst `nichts tun` — was der Bediener nicht
    ansieht, passiert auch nicht. Bei den eindeutigen ist es umgekehrt: die
    laufen durch, wenn niemand widerspricht.
    """
    if z.fall == FALL_AKTUALISIEREN:
        z.aktion = AKTION_AKTUALISIEREN
        z.ziel = z.bester.access if z.bester else None
        # Vorgehakt wird nur, was nichts verliert. Der Rest steht mit einem
        # Warnzeichen daneben und wartet auf eine ausdrückliche Entscheidung.
        z.uebernehmen = {
            f for f, (alt, neu) in abweichungen(z).items()
            if uebernahme_sinnvoll(f, alt, neu, z.ziel)
        } if UEBERNAHME_VORGABE else set()
    elif z.fall == FALL_NEU:
        z.aktion = AKTION_NEU
    else:
        z.aktion = AKTION_NICHTS


def auffaelligkeiten(lex: dict, satz: dict) -> list[str]:
    """Was an einem anzulegenden Satz auffällt.

    Kein Urteil, nur ein Fingerzeig: die Daten kommen aus einer über Jahre von
    Hand gepflegten Erfassung, und dort ist gelegentlich etwas ins falsche Feld
    gerutscht.

    Geprüft wird der **fertige Access-Satz** (`satz`), nicht der Lexware-Satz —
    denn der enthält bereits die Korrekturen des Bedieners. Sonst bliebe ein
    Hinweis stehen, nachdem man ihn gerade behoben hat.
    """
    def wert(feld):
        return str(satz.get(feld) or "").strip()

    probleme = []
    _, grund = merkmale(lex)
    if grund:
        probleme.append(grund)

    plz, land = wert("PLZ"), wert("Land")
    # Ausländische Postleitzahlen sind oft keine reinen Zahlen — „6721 CR"
    # ist ein gültiger niederländischer Code. Beanstandet wird deshalb nur,
    # was ohne Landangabe von der deutschen Form abweicht: dann ist entweder
    # die Straße ins PLZ-Feld geraten oder das Land fehlt.
    if not plz:
        probleme.append("keine PLZ")
    elif not land:
        if not plz.replace(" ", "").isdigit():
            probleme.append(f"PLZ ist keine Zahl ({plz}) und kein Land gesetzt")
        elif len(plz) != 5:
            probleme.append(f"PLZ hat {len(plz)} Stellen, kein Land gesetzt")

    if not wert("Straße"):
        probleme.append("keine Straße")
    elif not wert("Hausnummer"):
        probleme.append("keine Hausnummer")

    # Nicht beanstandet: eine Firma ohne Ansprechperson (Access führt 3 178
    # solcher Sätze, das ist der Normalfall) und eine Auslandsadresse (es gibt
    # schlicht Kunden in der Schweiz und den Niederlanden). Beides stand hier
    # einmal und färbte hunderte Zeilen ein, ohne dass etwas zu tun war —
    # eine Auffälligkeit, die überall aufleuchtet, ist keine.
    return probleme


# Felder, an denen man einen Kandidaten beurteilt. Reihenfolge wie man eine
# Adresse liest, nicht wie Access sie speichert.
VERGLEICHSFELDER = ["Name", "Vorname", "Institution", "PLZ", "Ort",
                    "Straße", "Hausnummer", "eMail"]

GLEICH, ANDERS, UNBEKANNT = "gleich", "anders", "unbekannt"


def gleichwertig(feld: str, a, b) -> bool:
    """Bedeuten zwei Werte dieses Feldes dasselbe?

    Je Feld die passende Normalisierung — „Kirchstraße" und „Kirchstr." sind
    dieselbe Straße und dürfen nicht als Abweichung erscheinen. Eine einzige
    Funktion dafür, damit Anzeige, Vorbelegung und Ausgabe nicht mit der Zeit
    auseinanderlaufen und dasselbe Feld mal als geändert gilt und mal nicht.
    """
    if feld == "Straße":
        return norm_strasse(a) == norm_strasse(b)
    if feld == "Hausnummer":
        return norm_hausnr(a) == norm_hausnr(b)
    if feld == "eMail":
        return norm_mail(a) == norm_mail(b)
    return norm_text(a) == norm_text(b)


def vergleiche(lex: dict, acc: dict) -> list[tuple]:
    """Feld für Feld: (Feld, Lexware-Wert, Access-Wert, Befund).

    `UNBEKANNT` heißt, dass eine Seite nichts weiß — das ist ausdrücklich
    etwas anderes als ein Widerspruch und soll auch anders aussehen. Sonst
    liest sich eine fehlende E-Mail wie eine falsche.
    """
    werte = lexware_werte(lex)
    ergebnis = []
    for feld in VERGLEICHSFELDER:
        l = str(werte.get(feld) or "").strip()
        a = str(acc.get(feld) or "").strip()
        if not l or not a:
            befund = UNBEKANNT
        else:
            befund = GLEICH if gleichwertig(feld, l, a) else ANDERS
        ergebnis.append((feld, l, a, befund))
    return ergebnis


def wert_geht_verloren(alt, neu) -> bool:
    """Verliert der Access-Wert Angaben, wenn `neu` ihn ersetzt?

    Nicht dasselbe wie „ist kürzer": „Mannheim" statt „Neckargemünd" ist ein
    Umzug, kein Verlust, auch wenn es vier Zeichen weniger sind. Verlust ist
    es, wenn der neue Wert im alten bereits steckt oder ihn nur abkürzt:

        Vaihingen                    steckt in  Vaihingen/Enz
        07224 40133                  steckt in  zzz_keine Werbeanrufe! 07224 40133
        Württemb. Landesbibliothek   kürzt ab   Württembergische Landesbibliothek

    Solche Übernahmen löschen etwas, das jemand einmal absichtlich erfasst hat.
    """
    a, n = norm_text(alt), norm_text(neu)
    if not a or not n or a == n:
        return False
    if n in a:
        return True
    worte_a, worte_n = a.split(), n.split()
    return (len(worte_n) <= len(worte_a)
            and all(any(wa.startswith(wn) for wa in worte_a)
                    for wn in worte_n))


def uebernahme_sinnvoll(feld: str, alt, neu, ziel: dict) -> bool:
    """Soll diese Übernahme überhaupt vorgeschlagen werden?

    Der Grundsatz: eine Übernahme darf die Anschrift ändern, aber nicht
    ärmer machen. Was Angaben verliert, wird angezeigt und gekennzeichnet,
    aber nicht vorgehakt — der Bediener kann es trotzdem anhaken.
    """
    if wert_geht_verloren(alt, neu):
        return False
    # „Damen und Herren" ist die Anrede für eine Einrichtung ohne
    # Ansprechperson. Kennt Access dort einen Namen, ist die persönliche
    # Anrede die bessere, und Lexware weiß es bloß nicht besser.
    if (feld == "Anrede" and norm_text(neu) == ANREDE_SAMMEL and alt
            and (str(ziel.get("Name") or "").strip()
                 or str(ziel.get("Vorname") or "").strip())):
        return False
    return True


def verschmelze_saetze(saetze: list[dict]) -> dict:
    """Mehrere Access-Sätze zu einem verschmelzen.

    Je Feld gewinnt der erste nicht leere Wert — die Reihenfolge bestimmt also
    der Aufrufer, indem er den zu behaltenden Satz nach vorn stellt. Bei
    E-Mail und Telefon werden stattdessen ALLE verschiedenen Werte behalten
    und mit Komma verbunden: zwei Anschriften desselben Menschen haben oft
    zwei Adressen, und keine davon ist falsch.
    """
    if not saetze:
        return {}
    ergebnis = dict(saetze[0])
    for feld in set().union(*(s.keys() for s in saetze)):
        werte = []
        for satz in saetze:
            for teil in str(satz.get(feld) or "").split(TRENNER_MEHRWERTIG):
                teil = teil.strip()
                if teil and teil not in werte:
                    werte.append(teil)
        if not werte:
            continue
        ergebnis[feld] = (TRENNER_MEHRWERTIG.join(werte)
                          if feld in MEHRWERTIG else werte[0])
    # Die ID des behaltenen Satzes bleibt, sonst zeigte der Import ins Leere.
    ergebnis["ID"] = saetze[0].get("ID")
    return ergebnis


def geaenderte_felder(z: Zuordnung, zeile: dict) -> list[str]:
    """Felder, die der Import in Access tatsächlich verändern wird.

    Nicht dasselbe wie `abweichungen`: das ist der Unterschied zwischen Access
    und Lexware, also die Auswahl, die dem Bediener vorgelegt wird. Hier zählt
    das Ergebnis — nach Häkchen und Handkorrekturen. Nur so kann die Liste
    zeigen, was der Lauf anrichtet, statt was er anrichten könnte.
    """
    if not z.ziel:
        return []
    return [f for f in AKTUALISIERBAR
            if not gleichwertig(f, zeile.get(f), z.ziel.get(f))]


def abweichungen(z: Zuordnung) -> dict[str, tuple]:
    """Felder, in denen Access und Lexware auseinandergehen: {Feld: (alt, neu)}.

    Leere Lexware-Felder zählen nicht als Abweichung — fehlende Angaben sind
    Nichtwissen, kein Widerspruch, und dürfen einen gepflegten Access-Wert
    nicht in Frage stellen.
    """
    if not z.ziel:
        return {}
    lex = lexware_werte(z.lexware)
    ergebnis = {}
    for feld in ADRESSFELDER:
        neu = str(lex.get(feld) or "").strip()
        alt = str(z.ziel.get(feld) or "").strip()
        if neu and not gleichwertig(feld, neu, alt):
            ergebnis[feld] = (alt, neu)
    return ergebnis


# ---------------------------------------------------------------------
# Feldabbildung Lexware -> Access
# ---------------------------------------------------------------------

ANREDE_SAMMEL = "damen und herren"

# Lexware führt akademische Titel im Vornamen ("Dr. Raimund Waibel"), Access
# hält sie getrennt: `Titel 2` für die akademischen (Dr., Prof. Dr.), `Titel 1`
# für Funktionen (Bürgermeister, Pfarrer, Dipl.-Ing.). In 18 366 Access-Sätzen
# steht kein einziges Mal ein Titel im Vornamensfeld — das ist gepflegt und
# soll so bleiben.
# Access trennt drei Arten von Titel, und sie gehören wirklich getrennt:
#   Titel 1  Berufsbezeichnung   Bürgermeister, Landrat, Dipl.-Ing.
#   Titel 2  akademisch, vorn    Dr., Prof. Dr., PD Dr.
#   Titel 3  nachgestellt        M.A., B.A., MdL
#
# Lexware kennt nur ein Vornamensfeld, in dem alles zusammen steht. Aus den
# 18 366 Access-Sätzen abgelesen: Titel 2 kommt 1 647-mal vor, Titel 3 98-mal,
# Titel 1 1 610-mal — letzteres aber pflegt der Verlag von Hand (1 237 davon
# tragen Kommunen-/Bürgermeister-Merkmale). In Lexware steht dazu nichts, also
# wird Titel 1 hier auch nicht geraten.
_TITEL2_TOKEN = re.compile(
    r"^(prof|dr|dres|pd|habil|h\.?c|phil|med|jur|rer|nat|ing)\.?$", re.I)
_TITEL1_VORN = re.compile(r"^dipl\.?[-\s]*(ing|kfm|päd|inf|biol)?\.?$", re.I)
_TITEL3_TOKEN = re.compile(
    r"^(m\.?\s?a|b\.?\s?a|m\.?\s?sc|b\.?\s?sc|ll\.?\s?m|mdl|mdb|m\.?ed)\.?$",
    re.I)


def trenne_titel(vorname) -> tuple[str, str, str, str]:
    """("Prof. Dr. Kurt M.A.") -> (Titel1, Titel2, Titel3, Vorname).

    „Dipl.-Ing." zählt als Berufsbezeichnung und damit zu Titel 1 — so führt
    Access es (24-mal), und so hat es der Verlag gemeint.
    """
    teile = str(vorname or "").split()

    titel1 = []
    while teile and _TITEL1_VORN.match(teile[0]):
        titel1.append(teile.pop(0))
    # „Dipl. Ing." steht als zwei Wörter da; das zweite gehört dazu.
    if titel1 and teile and teile[0].strip(".").lower() in (
            "ing", "kfm", "päd", "paed", "inf", "biol"):
        titel1.append(teile.pop(0))

    titel2 = []
    while teile and _TITEL2_TOKEN.match(teile[0].strip(".,")):
        titel2.append(teile.pop(0))

    titel3 = []
    while teile and _TITEL3_TOKEN.match(teile[-1].strip(".,")):
        titel3.insert(0, teile.pop())

    # „Dr," kommt in Lexware vor — beim Übernehmen gleich zu „Dr." glätten.
    def sauber(teile_):
        return " ".join(w.rstrip(",") + "." if w.rstrip(",").isalpha()
                        and not w.endswith(".") else w for w in teile_)

    return (sauber(titel1), sauber(titel2), sauber(titel3), " ".join(teile))


def bereinige_anrede(anrede, hat_firma: bool, hat_namen: bool) -> str:
    """Tippfehler glätten und die Sammelanrede richtig setzen.

    „Damen und Herren" gehört an Einrichtungen ohne Ansprechperson — 3 172 von
    3 178 solcher Access-Sätze führen sie. Sobald aber ein Name dasteht, ist
    sie falsch: dann wird die Person angeschrieben, nicht das Haus. Beide
    Richtungen werden hier hergestellt, damit niemand sie von Hand nachziehen
    muss.
    """
    s = str(anrede or "").strip()
    ersatz = {"herrn": "Herr", "herr": "Herr", "fraz": "Frau", "frau": "Frau"}
    s = ersatz.get(s.lower(), s)

    if hat_namen:
        # Mit Namen keine Sammelanrede — lieber gar keine als eine falsche.
        return "" if norm_text(s) == ANREDE_SAMMEL else s
    if not s and hat_firma:
        return "Damen und Herren"
    return s


# Merkmalsfelder aus Kundengruppe/Branche. ACHTUNG: aus den Häufigkeiten des
# Altbestands abgeleitet, nicht aus einer Vorgabe des Verlags — vor dem ersten
# scharfen Lauf gegenprüfen. Unbekannte Kombinationen landen auf "K" und
# werden in der GUI als "Merkmal prüfen" markiert.
MERKMAL_GRUPPE = {
    "Endkunden": "K",
    "Wiederverkäufer": "BUHA",
    "Presse": "P",
    "HA / A": "VIP",
    "VIP": "VIP",
    "Anzeigenkunden": "K",
    "Lieferanten": "K",
}
MERKMAL_BRANCHE = {
    "Autoren": {"Autor": "A"},
    "Mitautor/in": {"Autor": "MA"},
    "Herausgeber": {"Autor": "HA"},
    "Zeitschrift": {"Presse/ZS": "ZS"},
    "Tages/Wochenpresse": {"Presse/ZS": "Presse"},
    "freier Journalist": {"Presse/ZS": "Presse"},
    "Stadt / Gemeinde": {"Kommunen": "Komm"},
    "Verein": {"Verein": "Verein"},
    "Schulen": {"Schule": "Schule"},
    "Universität": {"Schule": "Schule"},
    "Archiv": {"Archiv": "Archiv"},
    "Bibliotheken": {"Archiv": "Archiv"},
}

# Ersatzweise aus der Branche, wenn die Kundengruppe fehlt. Die Auftragsliste
# führt `Kd.Gr.` nämlich oft nicht, wohl aber `Branche` — und wer unter
# "Buha" oder "Großhandel" bestellt, ist ganz sicher kein Endkunde. Ohne diese
# Ableitung landeten Buchhändler und Autoren pauschal auf "K".
MERKMAL_BRANCHE_GRUPPE = {
    "Buha": "BUHA",
    "Großhandel": "BUHA",
    "Onlinehandel": "BUHA",
    "Autoren": "VIP",
    "Mitautor/in": "VIP",
    "Herausgeber": "VIP",
    "Zeitschrift": "P",
    "Tages/Wochenpresse": "P",
    "freier Journalist": "P",
    "Bibliotheken": "Bib",
    "Archiv": "Bib",
    "Universität": "Wiss",
    "Schulen": "Wiss",
    "Museen": "Wiss",
    "Privat": "K",
    "Stadt / Gemeinde": "K",
    "Behörde": "K",
    "Verein": "K",
    "Unternehmen": "K",
    "Verlag": "K",
    "Gastronomie": "K",
}


def merkmale(lex: dict) -> tuple[dict, str | None]:
    """Merkmalsfelder für einen neuen Access-Satz.

    Zweiter Rückgabewert ist ein Grund, falls `VIP/W/K/X` geraten werden
    musste — sonst None. Der Grund landet als Auffälligkeit in der Liste, und
    er soll benennen, WAS fehlt: eine nicht exportierte Spalte verlangt eine
    andere Reaktion als eine unbekannte Kundengruppe.
    """
    gruppe = str(lex.get("Kundengruppe") or "").strip()
    branche = str(lex.get("Branche") or "").strip()

    grund = None
    if gruppe in MERKMAL_GRUPPE:
        haupt = MERKMAL_GRUPPE[gruppe]
    elif branche in MERKMAL_BRANCHE_GRUPPE:
        haupt = MERKMAL_BRANCHE_GRUPPE[branche]
        if gruppe:
            grund = f"Kundengruppe „{gruppe}“ unbekannt, aus Branche abgeleitet"
    else:
        haupt = "K"
        if gruppe:
            grund = f"Kundengruppe „{gruppe}“ unbekannt, „K“ angenommen"
        elif branche:
            grund = f"Branche „{branche}“ unbekannt, „K“ angenommen"
        else:
            grund = "weder Kundengruppe noch Branche im Export, „K“ angenommen"

    werte = {"VIP/W/K/X": haupt}
    werte.update(MERKMAL_BRANCHE.get(branche, {}))
    return werte, grund


# Reihenfolge für die Eingabemaske: erst das, was man beim Durchsehen
# tatsächlich anfasst, gruppiert wie eine Adresse gelesen wird. Der Rest der
# 50 Access-Spalten (Regionen, Fachgebiete, Naturwiss/Geisteswiss, Newsletter,
# Bemerkungen …) kommt darunter — die sind für einen frischen Lexware-Satz so
# gut wie immer leer und würden oben nur den Blick verstellen.
FELDER_GAENGIG = [
    "ID",
    "Anrede", "Titel 2", "Vorname", "Name",
    "Institution", "Abteilung",
    "Straße", "Hausnummer", "PLZ", "Ort", "Land",
    "eMail", "Telefon1", "Telefax",
    "Bestelldatum", "VIP/W/K/X", "Autor", "Presse/ZS",
    "Quelle", "erfaßt/geprüft am", "Lexware-Kd-Nr",
]


def ordne_felder(satz: dict) -> tuple[list, list]:
    """Teilt einen Satz in (gängige Felder, übrige) — jeweils (Feld, Wert)."""
    gaengig = [(f, satz[f]) for f in FELDER_GAENGIG if f in satz]
    bekannt = set(FELDER_GAENGIG)
    rest = [(f, w) for f, w in satz.items() if f not in bekannt]
    return gaengig, rest


# Felder, die mit Lexware verglichen und zur Übernahme angeboten werden.
ADRESSFELDER = [
    "Anrede", "Name", "Vorname", "Institution", "Abteilung",
    "PLZ", "Ort", "Straße", "Hausnummer", "Land",
    "Telefon1", "Telefon2", "Telefon3", "Telefax", "eMail",
]

# Einordnende Felder. Sie werden von einer Aktualisierung mitgeschrieben —
# damit sich eine falsche Kategorie beim Durchsehen richten lässt — aber
# NIEMALS aus Lexware vorgeschlagen: die dortige Kundengruppe ist gröber als
# die über Jahre gepflegte Einordnung in Access, und `merkmale()` rät sie
# obendrein. Ändern kann sie nur der Bediener von Hand.
MERKMALFELDER = ["Titel 2", "VIP/W/K/X", "Autor", "Presse/ZS"]

# Was eine Aktualisierung insgesamt schreibt.
AKTUALISIERBAR = ADRESSFELDER + MERKMALFELDER
# Diese setzt das Tool immer selbst, sie sind nicht verhandelbar.
STAMMFELDER = ["Bestelldatum", "Lexware-Kd-Nr", "Quelle", "erfaßt/geprüft am"]


def lexware_werte(lex: dict) -> dict:
    """Der Lexware-Satz, übersetzt in Access-Feldnamen."""
    titel1, titel2, titel3, vorname = trenne_titel(lex.get("Vorname"))
    firma = str(lex.get("Firma") or "").strip()
    # Nachgestellte Grade hängen auch mal am Nachnamen („Müller M.A.").
    _, _, titel3_name, name = trenne_titel(lex.get("Name"))
    titel3 = " ".join(x for x in (titel3, titel3_name) if x)
    strasse, hausnr = teile_strasse(lex.get("Straße"), lex.get("Haus Nr."))

    # Lexware kennt nur ein „Zusatz"-Feld, Access unterscheidet `Abteilung`
    # (Fachbereich im Haus) und `c/o` (Empfang über jemand anderen). Ein
    # „c/o …" gehört in die Zustellzeile, nicht in die Abteilung — sonst
    # steht es im Brief an der falschen Stelle.
    # Die drei Freifelder aus Lexware sind Notizen derselben Art, die in
    # Access unter `Bemerkungen` stehen („Freie Journalistin", „MA: Räume des
    # Glaubens"). Sie werden zusammengefasst, damit nichts davon verlorengeht.
    freifelder = TRENNER_BEMERKUNG.join(
        w for w in (str(lex.get(f) or "").strip()
                    for f in ("Freifeld1", "Freifeld2", "Freifeld3")) if w)

    zusatz = str(lex.get("Zusatz") or "").strip()
    care_of = zusatz if norm_text(zusatz).startswith("c o") else ""
    abteilung = "" if care_of else zusatz
    # Steht bei einer Firma kein Personenname, wandert die Firma nicht
    # zusätzlich ins Namensfeld — Access trennt das sauber.
    werte = {
        # `hat_namen` muss BEIDE Namensteile sehen: ein Satz mit Nachnamen,
        # aber ohne Vornamen ist trotzdem an eine Person gerichtet.
        "Anrede": bereinige_anrede(lex.get("Anrede"), bool(firma),
                                   bool(vorname or name)),
        "Name": name if name != firma else "",
        "Vorname": vorname,
        # Titel 1 bleibt leer: die Berufsbezeichnung (Bürgermeister, Landrat)
        # pflegt der Verlag in Access von Hand, Lexware weiß davon nichts.
        # Beim Aktualisieren bleibt der vorhandene Access-Wert deshalb stehen.
        "Titel 1": titel1,
        "Titel 2": titel2,
        "Titel 3": titel3,
        "Institution": firma,
        "Abteilung": abteilung,
        "c/o": care_of,
        "Bemerkungen": freifelder,
        "PLZ": str(lex.get("Plz") or "").strip(),
        "Ort": str(lex.get("Ort") or "").strip(),
        "Straße": strasse,
        "Hausnummer": hausnr,
        "Land": str(lex.get("Land") or "").strip(),
        "Telefon1": str(lex.get("Tel1") or "").strip(),
        "Telefon2": str(lex.get("Mobil") or "").strip(),
        "Telefon3": str(lex.get("Tel3") or "").strip(),
        "Telefax": str(lex.get("Fax") or "").strip(),
        "eMail": str(lex.get("E-Mail") or "").strip(),
    }
    return werte


# Felder, die in Access eine Zahl sind. Aus der Eingabemaske kommt jeder Wert
# als Text — und aus entscheidungen.json ebenso, dort wird alles als
# Zeichenkette abgelegt. Ohne Umwandlung trifft die getippte „2026" später auf
# die 2025 aus Access, und der Vergleich beim Zusammenführen scheitert mit
# „'>' not supported between instances of str and int".
ZAHLENFELDER = ("Bestelldatum",)


def typisiere(feld: str, wert):
    """Einen Wert aus der Eingabemaske in den Typ bringen, den Access führt."""
    if feld not in ZAHLENFELDER:
        return wert
    text = str(wert if wert is not None else "").strip()
    if not text:
        return None
    treffer = re.search(r"\d{4}", text)
    return int(treffer.group(0)) if treffer else None


def handkorrekturen(z: Zuordnung) -> dict:
    """Die Änderungen des Bedieners, typrichtig."""
    return {f: typisiere(f, w) for f, w in z.aenderungen.items()}


def baue_neuen_satz(z: Zuordnung, spalten: list[str], laufjahr: int,
                    heute) -> dict:
    """Ein anlagefertiger Access-Satz, Spalte für Spalte."""
    satz = {s: None for s in spalten}
    satz.pop("ID", None)   # AutoWert vergibt Access selbst

    satz.update(lexware_werte(z.lexware))
    merk, _ = merkmale(z.lexware)
    satz.update(merk)

    satz["Quelle"] = f"Lexware Kunden {laufjahr}"
    satz["erfaßt/geprüft am"] = heute
    satz["Bestelldatum"] = z.bestelljahr        # None bleibt leer (Freiexemplar)
    satz["Lexware-Kd-Nr"] = str(z.lexware.get("Kd.-Nr") or "").strip()

    satz.update(handkorrekturen(z))             # Handkorrekturen zuletzt
    return satz


def baue_aktualisierung(z: Zuordnung, laufjahr: int, heute) -> dict:
    """Zeile für aktualisieren.xlsx: alle Felder tragen bereits den Endwert.

    Dadurch bleibt die UPDATE-Abfrage in Access stumpf, und die Datei ist
    lesbar — man sieht der Zeile an, wie der Datensatz hinterher aussieht.
    """
    acc = z.ziel or {}
    lex = lexware_werte(z.lexware)

    # Vom vollständigen Access-Satz ausgehen, nicht von einer Auswahl: seit
    # die Ausgabe die ganze Tabelle schreibt, gibt es keinen Grund mehr,
    # Felder auszusparen — und die Eingabemaske zeigt in beiden Reitern
    # dieselben Felder. Unberührtes trägt weiterhin den Access-Wert.
    zeile = dict(acc)

    for feld in ADRESSFELDER:
        if feld in z.uebernehmen:
            zeile[feld] = lex.get(feld)

    # Bemerkungen werden ANGEHÄNGT, nie ersetzt: in Access steht dort oft
    # etwas, das jemand vor Jahren notiert hat, und die Lexware-Freifelder
    # sagen etwas anderes. Die Prüfung auf „steckt schon drin" macht das
    # wiederholbar — ein zweiter Lauf verdoppelt den Text nicht.
    notiz = str(lex.get("Bemerkungen") or "").strip()
    if notiz:
        vorhanden = str(acc.get("Bemerkungen") or "").strip()
        if not vorhanden:
            zeile["Bemerkungen"] = notiz
        elif norm_text(notiz) not in norm_text(vorhanden):
            zeile["Bemerkungen"] = vorhanden + TRENNER_BEMERKUNG + notiz

    alt = acc.get("Bestelldatum")
    alt = alt if isinstance(alt, int) else None
    # Nur anheben, nie senken: ein älteres Bestelljahr aus einer Nachlieferung
    # darf ein neueres in Access nicht überschreiben.
    jahre = [x for x in (alt, z.bestelljahr) if isinstance(x, int)]
    zeile["Bestelldatum"] = max(jahre) if jahre else None
    zeile["Lexware-Kd-Nr"] = str(z.lexware.get("Kd.-Nr") or "").strip()
    zeile["Quelle"] = f"Lexware Kunden {laufjahr}"
    zeile["erfaßt/geprüft am"] = heute

    zeile.update(handkorrekturen(z))
    return zeile


def gleiche_alle_ab(kunden: list[dict], access: list[dict],
                    lage: Auftragslage) -> list[Zuordnung]:
    """Kompletter Durchlauf: je Lexware-Kunde eine Zuordnung."""
    indizes = baue_indizes(access)
    ergebnis = []
    for lex in kunden:
        kdnr = str(lex.get("Kd.-Nr") or "").strip()
        kandidaten = finde_zuordnung(lex, indizes)
        ergebnis.append(stufe_ein(
            lex, kandidaten,
            lage.kaufjahr.get(kdnr),
            kdnr in lage.hat_beleg,
        ))

    # Zeigen zwei Lexware-Sätze auf denselben Access-Satz, ist mindestens einer
    # falsch — oder es ist der gewollte Zweitadressen-Fall. Beide zur Hand.
    belegt = defaultdict(list)
    for z in ergebnis:
        if z.fall == FALL_AKTUALISIEREN and z.bester:
            belegt[id(z.bester.access)].append(z)
    for gruppe in belegt.values():
        if len(gruppe) > 1:
            for z in gruppe:
                z.fall = FALL_UNKLAR
                # Auch als Ursprungsfall festhalten, sonst hebt ein
                # Zurücksetzen diesen Schutz wieder auf und die Kollision wäre
                # still zurück.
                z.fall_urspruenglich = FALL_UNKLAR
                # Die anderen Kundennummern benennen. „Mehrere Lexware-Kunden"
                # allein half niemandem: man sah nicht, welche gemeint sind,
                # und erst recht nicht, dass das Werkzeug sie beim Schreiben
                # ohnehin zusammenführt.
                andere = [str(a.lexware.get("Kd.-Nr") or "")
                          for a in gruppe if a is not z]
                z.unklar_grund.append(
                    f"Kd.-Nr. {', '.join(andere)} "
                    f"{'zeigt' if len(andere) == 1 else 'zeigen'} auf "
                    f"denselben Access-Satz — offenbar in Lexware doppelt "
                    f"angelegt. Bei „aktualisieren\" werden sie zu EINER Zeile "
                    f"zusammengeführt, mit allen Kundennummern im Feld "
                    f"Lexware-Kd-Nr.")
                _vorbelegen(z)
    return ergebnis


# ---------------------------------------------------------------------
# Ausgabe
# ---------------------------------------------------------------------

def _schreibe_blatt(pfad, spalten: list[str], zeilen: list[dict],
                    titel: str = "Tabelle1") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titel
    ws.append(spalten)
    for zeile in zeilen:
        ws.append([zeile.get(s) for s in spalten])
    ws.freeze_panes = "A2"
    wb.save(pfad)


# Mehrere Lexware-Nummern in einem Access-Feld. Text(20) reicht für zwei
# sechsstellige Nummern samt Trenner.
TRENNER_KDNR = ", "


def zerlege_kdnr(wert) -> list[str]:
    """Ein `Lexware-Kd-Nr`-Feld in einzelne Nummern zerlegen."""
    return [t.strip() for t in
            str(wert or "").replace(";", ",").split(",") if t.strip()]


def _verschmelze(a: dict, az: Zuordnung, b: dict, bz: Zuordnung,
                 ziel: dict) -> dict:
    """Zwei Aktualisierungen desselben Access-Satzes zu einer machen."""
    neu = dict(a)

    # isinstance statt "is not None": ein Fremdtyp darf hier nicht den ganzen
    # Lauf abbrechen, nachdem jemand tausend Fälle durchgesehen hat.
    jahre = [typisiere("Bestelldatum", j)
             for j in (a.get("Bestelldatum"), b.get("Bestelldatum"))]
    jahre = [j for j in jahre if isinstance(j, int)]
    neu["Bestelldatum"] = max(jahre) if jahre else None

    # Beide Kundennummern behalten. Nächstes Jahr findet Stufe 0 den Satz
    # dann über jede von beiden — und man sieht in Access, dass Lexware
    # denselben Menschen zweimal führt.
    nummern = []
    for wert in (a.get("Lexware-Kd-Nr"), b.get("Lexware-Kd-Nr")):
        for nr in zerlege_kdnr(wert):
            if nr not in nummern:
                nummern.append(nr)
    neu["Lexware-Kd-Nr"] = TRENNER_KDNR.join(sorted(nummern))

    # Adressfelder: wer den Access-Wert überhaupt ändert, gewinnt gegen wen,
    # der ihn stehen lässt. Ändern beide dasselbe Feld, gewinnt die jüngere
    # Bestellung — die aktuellere Anschrift.
    juenger = (bz.bestelljahr or 0) > (az.bestelljahr or 0)
    for feld in AKTUALISIERBAR:
        alt = norm_text(ziel.get(feld))
        a_aendert = norm_text(a.get(feld)) != alt
        b_aendert = norm_text(b.get(feld)) != alt
        if b_aendert and (not a_aendert or juenger):
            neu[feld] = b.get(feld)
    return neu


def verschmolzene_aktualisierungen(zuordnungen: list[Zuordnung], laufjahr: int,
                                   heute) -> tuple[list[dict], list[str]]:
    """Je Access-Satz genau EINE Zeile. Gibt (Zeilen, Anmerkungen) zurück.

    Zwei Lexware-Kundennummern können auf denselben Access-Satz zeigen —
    derselbe Mensch, zweimal in Lexware angelegt. Ohne Zusammenführung stünde
    die ID zweimal in aktualisieren.xlsx, und Access' UPDATE nähme davon
    irgendeine. Das darf nicht dem Zufall überlassen bleiben.
    """
    nach_id: dict = {}
    anmerkungen = []
    for z in zuordnungen:
        if z.aktion != AKTION_AKTUALISIEREN or not z.ziel:
            continue
        zeile = baue_aktualisierung(z, laufjahr, heute)
        schluessel = z.ziel.get("ID")
        if schluessel not in nach_id:
            nach_id[schluessel] = (zeile, z)
            continue
        alt_zeile, alt_z = nach_id[schluessel]
        nach_id[schluessel] = (
            _verschmelze(alt_zeile, alt_z, zeile, z, z.ziel), z)
        anmerkungen.append(
            f'Access-ID {schluessel}: Kd.-Nr. '
            f'{alt_z.lexware.get("Kd.-Nr")} und {z.lexware.get("Kd.-Nr")} '
            f'zeigen auf denselben Satz — zu einer Zeile zusammengeführt, '
            f'beide Nummern eingetragen.')
    return [zeile for zeile, _ in nach_id.values()], anmerkungen


def schreibe_neu_anlegen(pfad, zuordnungen: list[Zuordnung],
                         access_spalten: list[str], laufjahr: int, heute):
    """Spaltenreihenfolge exakt wie im Access-Export, nur ohne `ID` und mit
    `Lexware-Kd-Nr` — sonst scheitert der Anfüge-Import wortlos."""
    spalten = [s for s in access_spalten if s != "ID"]
    if "Lexware-Kd-Nr" not in spalten:
        spalten.append("Lexware-Kd-Nr")
    zeilen = [baue_neuen_satz(z, spalten, laufjahr, heute)
              for z in zuordnungen if z.aktion == AKTION_NEU]
    _schreibe_blatt(pfad, spalten, zeilen)
    return len(zeilen)


def schreibe_aktualisieren(pfad, zuordnungen: list[Zuordnung],
                           laufjahr: int, heute):
    spalten = ["ID"] + STAMMFELDER + AKTUALISIERBAR
    zeilen, _ = verschmolzene_aktualisierungen(zuordnungen, laufjahr, heute)
    _schreibe_blatt(pfad, spalten, zeilen)
    return len(zeilen)


def schreibe_gesamttabelle(pfad, access: list[dict], zuordnungen: list[Zuordnung],
                           spalten: list[str], laufjahr: int, heute,
                           entfernen=None, markieren=None) -> int:
    """Die komplette Kundentabelle, fertig aktualisiert und ergänzt.

    Gedacht für den Weg ohne SQL: in Access die Tabelle leeren und diese eine
    Datei anhängen. Das erspart Importtabelle und UPDATE-Abfrage — Dinge, die
    einmal im Jahr niemand parat hat.

    Der Preis: die Datei ersetzt den gesamten Bestand. Wer hier fehlt, ist
    hinterher weg. Deshalb wird sie NUR geschrieben, wenn der Access-Export
    nachweislich die volle Tabelle ist und nicht die Mailing-Abfrage — siehe
    `pruefe_access_vollstaendigkeit`.
    """
    # Über dieselbe Zusammenführung wie aktualisieren.xlsx, sonst gewönne bei
    # zwei Kundennummern auf einem Access-Satz still die zuletzt einsortierte.
    zusammengefasst, _ = verschmolzene_aktualisierungen(
        zuordnungen, laufjahr, heute)
    aenderungen = {zeile.get("ID"): zeile for zeile in zusammengefasst}

    # Zusammengelegte Sätze fallen ersatzlos weg — ihre Angaben stecken im
    # behaltenen Satz. Nur über die Gesamttabelle ist ein Zusammenlegen
    # überhaupt möglich; ein Anfüge-Import kann nichts entfernen.
    aufgeloest = {i for z in zuordnungen for i in z.aufgeloest_in}
    aufgeloest |= set(entfernen or ())
    markieren = dict(markieren or {})

    zeilen = []
    for satz in access:
        if satz.get("ID") in aufgeloest:
            continue
        zeile = dict(satz)
        # Statt zu entfernen kann man auch nur kennzeichnen — so hält es der
        # Verlag bisher: 81 Sätze tragen „verzogen", „verstorben" oder
        # „Mailing kam zurück" und bleiben trotzdem in der Tabelle. Das
        # bewahrt die Vorgeschichte, die ein Löschen wegwirft.
        vermerk = markieren.get(satz.get("ID"))
        if vermerk:
            zeile["Datensatz gelöscht"] = vermerk
        neu = aenderungen.get(satz.get("ID"))
        if neu:
            for feld, wert in neu.items():
                if feld != "ID":          # der Schlüssel bleibt, wie er ist
                    zeile[feld] = wert
        zeilen.append(zeile)

    # Neuanlagen ans Ende. Sie tragen keine ID — die vergibt Access selbst.
    zeilen.extend(baue_neuen_satz(z, spalten, laufjahr, heute)
                  for z in zuordnungen if z.aktion == AKTION_NEU)

    ausgabe = list(spalten)
    if "Lexware-Kd-Nr" not in ausgabe:
        ausgabe.append("Lexware-Kd-Nr")
    _schreibe_blatt(pfad, ausgabe, zeilen)
    return len(zeilen)


def schreibe_protokoll(pfad, zuordnungen: list[Zuordnung], befund=None):
    """Jede Entscheidung mit Punktzahl und Begründung — die Spur, an der man
    nächstes Jahr nachvollzieht, warum etwas passiert ist."""
    spalten = ["Kd.-Nr", "Lexware", "Fall", "Aktion", "Bestelljahr",
               "Access-ID", "Punkte", "Begründung", "übernommene Felder",
               "Hinweise", "warum unklar", "Access-Dubletten",
               "weitere Kandidaten"]
    zeilen = []
    for z in zuordnungen:
        lex = z.lexware
        name = " ".join(x for x in (str(lex.get("Vorname") or ""),
                                    str(lex.get("Name") or ""),
                                    str(lex.get("Firma") or "")) if x)
        weitere = "; ".join(
            f'{k.punkte:.0f} ID {k.access.get("ID")}' for k in z.kandidaten[1:4])
        zeilen.append({
            "Kd.-Nr": str(lex.get("Kd.-Nr") or ""),
            "Lexware": f'{name}, {lex.get("Plz") or ""} {lex.get("Ort") or ""}',
            "Fall": z.fall,
            "Aktion": z.aktion,
            "Bestelljahr": z.bestelljahr,
            "Access-ID": (z.ziel or {}).get("ID"),
            "Punkte": z.bester.punkte if z.bester else None,
            "Begründung": z.bester.begruendung if z.bester else "",
            "übernommene Felder": ", ".join(sorted(z.uebernehmen)),
            "Hinweise": " | ".join(z.hinweise),
            "warum unklar": " | ".join(z.unklar_grund),
            # Gleichlautende Access-Sätze, die zusammengefasst wurden. Gehört
            # ins Protokoll, weil sie sonst niemand je zu sehen bekommt — und
            # weil der Verlag sie in Access aufräumen könnte.
            "Access-Dubletten": ", ".join(
                str(i) for k in z.kandidaten for i in k.dubletten),
            "weitere Kandidaten": weitere,
        })
    _schreibe_blatt(pfad, spalten, zeilen)

    if befund is not None and befund.verdaechtig:
        # Die Warnung gehört mit ins Protokoll, sonst ist beim Nachlesen nicht
        # mehr erkennbar, dass der Lauf auf lückenhaften Eingaben beruhte.
        wb = openpyxl.load_workbook(pfad)
        ws = wb.create_sheet("Warnung")
        ws["A1"] = "Plausibilitätsprüfung der Eingaben"
        ws["A2"] = befund.text
        ws.column_dimensions["A"].width = 120
        ws["A2"].alignment = openpyxl.styles.Alignment(wrapText=True)
        wb.save(pfad)
    return len(zeilen)


def schreibe_pflege(pfad, zuordnungen: list[Zuordnung],
                    access: list[dict]) -> str:
    """Was in Lexware und Access aufzuräumen wäre.

    Der Abgleich stolpert jedes Jahr über dieselben Nachlässigkeiten: eine
    Hausnummer im Straßenfeld, ein doppelt angelegter Kunde, eine PLZ, die
    keine ist. Behoben wird das nicht hier, sondern in der Quelle — dafür muss
    aber jemand wissen, wo. Deshalb diese Liste; sie ändert nichts, sie zeigt.
    """
    def kdnr(z):
        return str(z.lexware.get("Kd.-Nr") or "")

    def wer(z):
        name = " ".join(x for x in (str(z.lexware.get("Vorname") or ""),
                                    str(z.lexware.get("Name") or "")) if x)
        firma = str(z.lexware.get("Firma") or "")
        return (f"{name} · {firma}" if name and firma else name or firma
                or "(ohne Namen)")

    abschnitte = []

    # --- Lexware: derselbe Mensch mehrfach angelegt
    doppelt = {}
    for z in zuordnungen:
        for grund in z.unklar_grund:
            if "denselben Access-Satz" in grund and z.ziel is not None:
                doppelt.setdefault(z.ziel.get("ID"), []).append(z)
            elif "denselben Access-Satz" in grund and z.bester:
                doppelt.setdefault(z.bester.access.get("ID"), []).append(z)
    if doppelt:
        zeilen = []
        for gruppe in doppelt.values():
            nummern = ", ".join(sorted(kdnr(z) for z in gruppe))
            zeilen.append(f"  {nummern:<24} {wer(gruppe[0])}")
        abschnitte.append((
            "LEXWARE: derselbe Kunde mehrfach angelegt",
            f"{len(doppelt)} Fälle. Beim naechsten Auftrag landet die Bestellung "
            f"mal unter der einen, mal unter der anderen Nummer.",
            sorted(zeilen)))

    # --- Lexware: Angaben im falschen Feld
    hausnr_in_strasse, plz_kaputt, ohne_strasse, ohne_hausnr = [], [], [], []
    for z in zuordnungen:
        roh_str = str(z.lexware.get("Straße") or "").strip()
        roh_nr = str(z.lexware.get("Haus Nr.") or "").strip()
        if not roh_nr and teile_strasse(roh_str)[1]:
            hausnr_in_strasse.append(f"  {kdnr(z):<10} „{roh_str}“   {wer(z)}")
        plz = str(z.lexware.get("Plz") or "").strip()
        land = str(z.lexware.get("Land") or "").strip()
        if not plz:
            plz_kaputt.append(f"  {kdnr(z):<10} PLZ fehlt   {wer(z)}")
        elif not land and not plz.replace(" ", "").isdigit():
            # Mit Landangabe wäre „6721 CR" völlig in Ordnung — ohne sie ist
            # entweder die Straße ins Feld geraten oder das Land vergessen.
            plz_kaputt.append(f"  {kdnr(z):<10} PLZ = „{plz}“, kein Land   "
                              f"{wer(z)}")
        if not roh_str:
            ohne_strasse.append(f"  {kdnr(z):<10} {wer(z)}")
        elif not roh_nr and not teile_strasse(roh_str)[1]:
            ohne_hausnr.append(f"  {kdnr(z):<10} „{roh_str}“   {wer(z)}")

    for titel, erklaerung, zeilen in (
            ("LEXWARE: Hausnummer steht im Straßenfeld",
             "Gehört ins Feld „Haus Nr.“. Das Werkzeug trennt es beim Lesen "
             "selbst, aber der Adressdruck tut das nicht.", hausnr_in_strasse),
            ("LEXWARE: PLZ fehlt oder ist keine Zahl",
             "Meist ist die Straße ins PLZ-Feld gerutscht.", plz_kaputt),
            ("LEXWARE: keine Straße erfasst", "", ohne_strasse),
            ("LEXWARE: keine Hausnummer erfasst",
             "Ohne sie lassen sich Nachbarn nicht unterscheiden.",
             ohne_hausnr)):
        if zeilen:
            abschnitte.append((titel, erklaerung, zeilen))

    # --- Lexware: Kundengruppe/Branche unbekannt
    unbekannt = []
    for z in zuordnungen:
        _, grund = merkmale(z.lexware)
        if grund:
            unbekannt.append(f"  {kdnr(z):<10} {grund}   {wer(z)}")
    if unbekannt:
        abschnitte.append((
            "LEXWARE: Kundengruppe oder Branche unbekannt",
            "Das Werkzeug hat „K“ angenommen. Entweder in Lexware nachtragen "
            "oder die Tabellen in core.py ergänzen.", unbekannt))

    # --- Access: gleichlautende Sätze
    gruppen = defaultdict(list)
    for satz in access:
        kennung = dubletten_kennung(satz)
        if any(kennung):
            gruppen[kennung].append(satz)
    access_doppelt = [g for g in gruppen.values() if len(g) > 1]
    if access_doppelt:
        zeilen = []
        for gruppe in access_doppelt:
            ids = ", ".join(str(s.get("ID")) for s in gruppe)
            name = " ".join(x for x in (str(gruppe[0].get("Vorname") or ""),
                                        str(gruppe[0].get("Name") or ""),
                                        str(gruppe[0].get("Institution") or ""))
                            if x)
            zeilen.append(f"  IDs {ids:<20} {name}")
        abschnitte.append((
            "ACCESS: gleichlautende Datensätze",
            f"{len(access_doppelt)} Gruppen. Sie bekommen jeder einen eigenen "
            f"Brief an dieselbe Anschrift.", sorted(zeilen)))

    grenze = 200
    teile = ["Was in den Quellsystemen aufzuräumen wäre",
             "=" * 44, "",
             "Diese Liste ändert nichts. Sie sammelt, worüber der Abgleich "
             "gestolpert ist —", "behoben wird es in Lexware bzw. Access, "
             "sonst steht es nächstes Jahr wieder da.", ""]
    for titel, erklaerung, zeilen in abschnitte:
        teile += ["", titel, "-" * len(titel)]
        if erklaerung:
            teile.append(erklaerung)
        teile.append(f"{len(zeilen)} Stück:")
        teile += zeilen[:grenze]
        if len(zeilen) > grenze:
            teile.append(f"  … und {len(zeilen) - grenze} weitere "
                         f"(hier gekürzt, vollständig in protokoll.xlsx)")
    if not abschnitte:
        teile.append("Nichts gefunden — die Quelldaten sind sauber.")

    Path(pfad).write_text("\n".join(teile) + "\n", encoding="utf-8")
    return f"{len(abschnitte)} Abschnitte"


def schreibe_zusammenlegen(pfad, zuordnungen: list[Zuordnung]):
    """Welche Access-Sätze in welchen aufgehen sollen.

    In kunden_komplett.xlsx sind sie bereits weg. Wer den Weg über die
    Einzeldateien geht, muss sie von Hand löschen — dafür ist diese Liste da.
    """
    spalten = ["behalten (ID)", "auflösen (ID)", "Kd.-Nr", "Lexware"]
    zeilen = []
    for z in zuordnungen:
        if not z.aufgeloest_in or not z.ziel:
            continue
        name = " ".join(x for x in (str(z.lexware.get("Vorname") or ""),
                                    str(z.lexware.get("Name") or ""),
                                    str(z.lexware.get("Firma") or "")) if x)
        for alt in z.aufgeloest_in:
            zeilen.append({"behalten (ID)": z.ziel.get("ID"),
                           "auflösen (ID)": alt,
                           "Kd.-Nr": str(z.lexware.get("Kd.-Nr") or ""),
                           "Lexware": name})
    _schreibe_blatt(pfad, spalten, zeilen)
    return len(zeilen)


def schreibe_zuordnung(pfad, zuordnungen: list[Zuordnung]):
    """Kd.-Nr. <-> Access-ID als Sicherung neben dem Access-Feld."""
    spalten = ["Kd.-Nr", "Access-ID", "Aktion"]
    zeilen = [{"Kd.-Nr": str(z.lexware.get("Kd.-Nr") or ""),
               "Access-ID": (z.ziel or {}).get("ID"),
               "Aktion": z.aktion}
              for z in zuordnungen
              if z.aktion == AKTION_AKTUALISIEREN and z.ziel]
    _schreibe_blatt(pfad, spalten, zeilen)
    return len(zeilen)


def schreibe_anleitung(pfad, tabelle: str = "Kunden", laufjahr: int = 0,
                       access_warnung=None) -> str:
    """Wie die eine Datei nach Access kommt."""
    warnung = ""
    if access_warnung:
        warnung = f"""
ACHTUNG — die geladene Access-Datei war unvollstaendig
=======================================================

{access_warnung}

Die Jahresdatei enthaelt daher nur die Saetze aus dieser Abfrage. Solange in
Schritt 1 eine KOPIE angelegt wird, ist nichts verloren: die bisherige Datei
bleibt unberuehrt, und man kann den Lauf mit dem richtigen Export wiederholen.
Zum Verschicken taugt das Ergebnis aber noch nicht.

"""

    text = f"""So kommt die Jahresdatei nach Access
=====================================
{warnung}
Es gibt genau EINE Datei zum Importieren: kunden_komplett.xlsx. Sie enthaelt
den vollstaendigen Kundenstamm, bereits aktualisiert, mit den neuen Kunden
hinten dran und den zusammengelegten Saetzen entfernt.

Der Kniff: an der bisherigen Datei wird nichts geaendert. Jedes Jahr entsteht
eine neue, die alte bleibt als Archiv liegen. Geht etwas schief, loescht man
die neue und faengt von vorn an.

  1. Die Access-Datei KOPIEREN und die Kopie nach dem Jahr benennen,
     z. B.  Adressen_{laufjahr}.accdb
     Ab hier wird nur noch in DIESER Kopie gearbeitet.

  2. Einmalig, falls noch nicht geschehen:
     In der Tabelle [{tabelle}] ein Feld  Lexware-Kd-Nr  anlegen (Text, 20).

  3. In der Kopie die Tabelle [{tabelle}] oeffnen, alle Datensaetze markieren
     (Strg+A) und loeschen. Die Tabelle ist danach leer — das ist so gewollt,
     und es betrifft nur die Kopie.

  4. Externe Daten -> Neue Datenquelle -> Aus Datei -> Excel
     kunden_komplett.xlsx waehlen
     "Eine Kopie der Datensaetze an die Tabelle anfuegen: {tabelle}"

  5. Datenbanktools -> Datenbank komprimieren und reparieren.
     Das setzt den AutoWert-Zaehler richtig, damit der naechste von Hand
     angelegte Datensatz keine schon vergebene ID bekommt.

  6. Stichprobe: ein paar bekannte Namen suchen und die Datensatzzahl unten
     links mit der Zeilenzahl in kunden_komplett.xlsx vergleichen.

  7. Von jetzt an ist  Adressen_{laufjahr}.accdb  die gueltige Datei.
     Die vorherige nicht loeschen — sie ist der Stand des Vorjahres.


WICHTIG: NICHT "in eine neue Tabelle importieren"
--------------------------------------------------
In Schritt 4 muss es "an die Tabelle anfuegen" heissen. Wer stattdessen eine
neue Tabelle importieren laesst, ueberlaesst Access das Raten der Feldtypen —
und dann wird aus der Postleitzahl 04103 die Zahl 4103. Betroffen waeren
225 Postleitzahlen und 9913 Telefonnummern, und die ID verloere ihre
AutoWert-Eigenschaft. Deshalb wird in die vorhandene Tabelle hinein angefuegt:
sie hat die Feldtypen seit 25 Jahren korrekt.


DIE ANDEREN DATEIEN
===================

protokoll.xlsx   jede Entscheidung mit Punktzahl und Begruendung — zum
                 Nachvollziehen, warum ein Satz so aussieht wie er aussieht.
                 Enthaelt die Mappe ein Blatt "Warnung", beruhte der Lauf auf
                 lueckenhaften Eingaben.

pflege.txt       was in Lexware und Access aufzuraeumen waere: doppelt
                 angelegte Kunden, Hausnummern im Strassenfeld, fehlende
                 Postleitzahlen. Das behebt dieses Werkzeug nicht — es zeigt
                 nur, wo es klemmt. Wer es in der Quelle richtet, hat es
                 naechstes Jahr nicht wieder.
"""
    Path(pfad).write_text(text, encoding="utf-8")
    return "geschrieben"


def schreibe_alles(ordner, zuordnungen: list[Zuordnung], access: list[dict],
                   access_spalten: list[str], laufjahr: int, heute,
                   befund=None, access_warnung=None,
                   entfernen=None, markieren=None) -> dict:
    """Die Ausgabe: eine Datei zum Importieren, drei zum Nachlesen.

    Es gab einmal fünf Excel-Dateien für zwei verschiedene Importwege. Das war
    für einen Vorgang, den jemand einmal im Jahr macht, eine Zumutung — und
    der zweite Weg brauchte obendrein eine SQL-Abfrage. Geblieben ist die
    Gesamttabelle: eine Datei, ein Import.
    """
    ordner = Path(ordner)
    ordner.mkdir(parents=True, exist_ok=True)

    # Was frühere Fassungen erzeugt haben, hier wegräumen. Sonst lägen im
    # Ordner fünf Excel-Dateien nebeneinander und niemand wüsste, welche die
    # gültige ist — der Grund, aus dem es nur noch eine gibt.
    for veraltet in ("neu_anlegen.xlsx", "aktualisieren.xlsx",
                     "zuordnung.xlsx", "zusammenlegen.xlsx",
                     "access_import.txt"):
        (ordner / veraltet).unlink(missing_ok=True)

    return {
        "kunden_komplett.xlsx": schreibe_gesamttabelle(
            ordner / "kunden_komplett.xlsx", access, zuordnungen,
            access_spalten, laufjahr, heute, entfernen, markieren),
        "anleitung.txt": schreibe_anleitung(
            ordner / "anleitung.txt", laufjahr=laufjahr,
            access_warnung=access_warnung),
        "pflege.txt": schreibe_pflege(
            ordner / "pflege.txt", zuordnungen, access),
        "protokoll.xlsx": schreibe_protokoll(
            ordner / "protokoll.xlsx", zuordnungen, befund),
    }
